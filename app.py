"""
app.py — Form 20 Backlog Dashboard
Flask backend with SQLite. Run: python app.py
"""

import csv
import io
import os
import json
import threading
import time
from datetime import datetime, timedelta

from dotenv import load_dotenv
load_dotenv()

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from functools import wraps
from authlib.integrations.flask_client import OAuth
from werkzeug.middleware.proxy_fix import ProxyFix

import gzip

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev_secret_key_change_me_in_production')
app.permanent_session_lifetime = timedelta(hours=1)

@app.after_request
def compress_response(response):
    accept_encoding = request.headers.get('Accept-Encoding', '')
    if 'gzip' not in accept_encoding.lower():
        return response
    if response.status_code < 200 or response.status_code >= 300 or 'Content-Encoding' in response.headers:
        return response
    content_type = response.headers.get('Content-Type', '')
    if not any(t in content_type for t in ['json', 'javascript', 'css', 'text', 'html', 'geojson']) and not request.path.endswith('.geojson'):
        return response
    if response.direct_passthrough:
        response.direct_passthrough = False
    data = response.get_data()
    if len(data) < 1024:
        return response
    compressed_data = gzip.compress(data)
    response.set_data(compressed_data)
    response.headers['Content-Encoding'] = 'gzip'
    response.headers['Content-Length'] = len(compressed_data)
    return response

oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=os.environ.get('GOOGLE_CLIENT_ID'),
    client_secret=os.environ.get('GOOGLE_CLIENT_SECRET'),
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'},
)

# ── Local-dev auth bypass ─────────────────────────────────────────────────────
# Google OAuth needs valid client credentials + a registered redirect URI, which
# aren't available in local development. When DISABLE_AUTH is set (auto-enabled for
# `python app.py` — see __main__), every request is treated as a signed-in dev user
# so the dashboard works on localhost without OAuth. Production (gunicorn) imports
# the module without __main__, so auth stays enforced unless DISABLE_AUTH is set.
DEV_USER = {
    'email':   'localdev@varaheanalytics.com',
    'name':    'Local Dev',
    'picture': '',
}

def auth_disabled():
    return os.environ.get('DISABLE_AUTH', '').strip().lower() in ('1', 'true', 'yes', 'on')


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if auth_disabled():
            return f(*args, **kwargs)
        if not session.get('user'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'data.db')
EXCEL_PATH = os.path.join(BASE_DIR, 'Form20 Backlog Tracker.xlsx')


# ── DB helpers ──────────────────────────────────────────────────────────────

class TursoCursorWrapper:
    def __init__(self, result_set):
        self.result_set = result_set
    
    def fetchall(self):
        cols = self.result_set.columns
        return [dict(zip(cols, list(r))) for r in self.result_set.rows]
        
    def fetchone(self):
        if not self.result_set.rows: return None
        cols = self.result_set.columns
        return dict(zip(cols, list(self.result_set.rows[0])))

class TursoConnectionWrapper:
    def __init__(self, client):
        self.client = client
    
    def cursor(self):
        return self
    
    def execute(self, query, params=()):
        result_set = self.client.execute(query, list(params))
        return TursoCursorWrapper(result_set)
        
    def close(self):
        self.client.close()
        
    def commit(self):
        pass

def get_db():
    import os
    turso_url = os.environ.get('TURSO_DATABASE_URL')
    turso_token = os.environ.get('TURSO_AUTH_TOKEN')
    
    if turso_url and turso_token:
        import libsql_client
        # Ensure HTTPS for robust connection
        turso_url = turso_url.replace('libsql://', 'https://')
        client = libsql_client.create_client_sync(url=turso_url, auth_token=turso_token)
        return TursoConnectionWrapper(client)
    else:
        import sqlite3
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn

def get_rds_db():
    import os
    import psycopg2
    
    # Credentials from AWS RDS
    db_host = os.environ.get('DB_HOST')
    db_port = os.environ.get('DB_PORT', '5432')
    db_user = os.environ.get('DB_USER')
    db_pass = os.environ.get('DB_PASSWORD')
    db_name = os.environ.get('DB_NAME')
    
    if not all([db_host, db_user, db_pass, db_name]):
        return None
        
    # connect_timeout + keepalives: without these a network blip leaves the
    # background refresh thread hung forever, and the single-flight lock then
    # blocks all future refreshes until the app restarts.
    conn = psycopg2.connect(
        host=db_host,
        port=db_port,
        user=db_user,
        password=db_pass,
        dbname=db_name,
        connect_timeout=120,
        keepalives=1,
        keepalives_idle=30,
        keepalives_interval=10,
        keepalives_count=3,
    )
    conn.autocommit = True
    return conn

# ── On-demand RDS cache: refresh on reload, NOT by background polling ─────────
# Dashboard data is served from a shared Redis cache for an instant response.
# A reload only touches RDS when the cached copy is older than CACHE_TTL_SECONDS,
# and even then the refresh runs once in the background, single-flight ACROSS ALL
# GUNICORN WORKERS (Redis SET NX EX lock) — concurrent reloads on any worker
# collapse into a single DB hit. If nobody opens the dashboard, RDS is never
# queried — load tracks real usage, not the clock.
#
# Cache lives in Redis on the EC2 host (docker-compose: service "redis"), NOT on
# local disk — nothing here writes cache files, so there's nothing to accidentally
# commit. If Redis is unreachable (e.g. local dev without the container), we
# degrade gracefully to a per-process in-memory dict — fine for a single dev
# process, but won't coordinate across workers (that's the Redis-only guarantee).
CACHE_TTL_SECONDS = 300   # 5 min — serve cache without hitting RDS within this window
REDIS_URL = os.environ.get('REDIS_URL', 'redis://127.0.0.1:6379/0')

# Redis is purely a CACHE — every helper below is wrapped so that ANY Redis
# failure (down, restarting, network blip, OOM-evicted key, etc.) degrades
# silently to the in-memory fallback or an empty result. A request must NEVER
# 500 because Redis had a bad moment.
_redis_client = None   # lazy singleton; False between failed connection attempts
_redis_warned = False
_redis_last_attempt = 0.0
REDIS_RETRY_INTERVAL = 30  # seconds before retrying a connection after a failure

_mem_cache_lock    = threading.Lock()
_mem_cache_data    = {}    # name -> last built payload   (in-memory fallback)
_last_refresh_ts   = {}    # name -> epoch seconds of last successful refresh (fallback)
_refresh_in_flight = set() # names whose background refresh is running         (fallback)


def get_redis():
    """Lazy singleton Redis client, or None if Redis is unreachable.
    Retries the connection every REDIS_RETRY_INTERVAL seconds (so if the Redis
    container starts after the app, or restarts, we self-heal without an app
    restart) — but never blocks/raises if it's down."""
    global _redis_client, _redis_warned, _redis_last_attempt
    if _redis_client not in (None, False):
        return _redis_client

    now = time.time()
    if _redis_client is False and (now - _redis_last_attempt) < REDIS_RETRY_INTERVAL:
        return None
    _redis_last_attempt = now

    try:
        import redis
        client = redis.Redis.from_url(
            REDIS_URL, decode_responses=True,
            socket_connect_timeout=0.2, socket_timeout=0.2,
        )
        client.ping()
        _redis_client = client
        if _redis_warned:
            print('[redis] connection restored — using shared cache again')
            _redis_warned = False
        return client
    except Exception as e:
        _redis_client = False
        if not _redis_warned:
            print(f'[redis] unavailable ({e}) — using in-memory cache (per-process fallback)')
            _redis_warned = True
        return None


def cache_get(name):
    """Return the cached payload for `name`, or None if not yet built.
    Never raises — any Redis error falls back to the in-memory copy."""
    r = get_redis()
    if r:
        try:
            raw = r.get(f'cache:{name}:data')
            if raw is not None:
                return json.loads(raw)
        except Exception as e:
            print(f'[redis] get failed for {name}: {e}')
    return _mem_cache_data.get(name)


def cache_set(name, data):
    """Persist `data` for `name` to Redis (shared across workers) AND keep an
    in-memory mirror — so a transient Redis outage in cache_get() still has
    last-known-good data to fall back to, instead of nothing."""
    _mem_cache_data[name] = data
    r = get_redis()
    if r:
        try:
            r.set(f'cache:{name}:data', json.dumps(data))
        except Exception as e:
            print(f'[redis] set failed for {name}: {e}')


def cache_is_fresh(name):
    """True if `name` was refreshed within the TTL (no DB hit needed)."""
    r = get_redis()
    if r:
        try:
            return r.exists(f'cache:{name}:fresh') == 1
        except Exception:
            return False
    return (time.time() - _last_refresh_ts.get(name, 0)) < CACHE_TTL_SECONDS


def trigger_refresh(name, builder):
    """Stale-while-revalidate, single-flight ACROSS ALL WORKERS via Redis. If
    `name`'s cache is stale and no refresh is already running anywhere, execute
    `builder()` once in a daemon thread. Returns immediately — the request is
    always served from the existing cache."""
    r = get_redis()
    if r:
        if cache_is_fresh(name):
            return
        lock_key = f'cache:{name}:lock'
        try:
            acquired = r.set(lock_key, '1', nx=True, ex=120)  # 2-min safety valve
        except Exception:
            acquired = False
        if not acquired:
            return  # another worker is already refreshing this cache

        def _run_redis():
            try:
                builder()
                r.set(f'cache:{name}:fresh', '1', ex=CACHE_TTL_SECONDS)
            except Exception as e:
                print(f'[cache:{name}] refresh failed: {e}')
            finally:
                try:
                    r.delete(lock_key)
                except Exception:
                    pass

        threading.Thread(target=_run_redis, daemon=True, name=f'refresh-{name}').start()
        return

    # ── Fallback: per-process lock/timestamp (single dev process, no Redis) ──
    with _mem_cache_lock:
        if cache_is_fresh(name) or name in _refresh_in_flight:
            return
        _refresh_in_flight.add(name)

    def _run_local():
        try:
            builder()
            with _mem_cache_lock:
                _last_refresh_ts[name] = time.time()
        except Exception as e:
            print(f'[cache:{name}] refresh failed: {e}')
        finally:
            with _mem_cache_lock:
                _refresh_in_flight.discard(name)

    threading.Thread(target=_run_local, daemon=True, name=f'refresh-{name}').start()


# Standard ECI state codes mapping
STATE_TO_ECI = {
    'AP': 'S01', 'AR': 'S02', 'AS': 'S03', 'BR': 'S04', 'GA': 'S05',
    'GJ': 'S06', 'HR': 'S07', 'HP': 'S08', 'KA': 'S10', 'KL': 'S11',
    'MP': 'S12', 'MH': 'S13', 'MN': 'S14', 'ML': 'S15', 'MZ': 'S16',
    'NL': 'S17', 'OR': 'S18', 'PB': 'S19', 'RJ': 'S20', 'SK': 'S21',
    'TN': 'S22', 'TR': 'S23', 'UP': 'S24', 'WB': 'S25', 'CG': 'S26',
    'JH': 'S27', 'UK': 'S28', 'TS': 'S29',
    'AN': 'U01', 'CH': 'U02', 'DD': 'U03', 'DN': 'U03', 'DL': 'U05',
    'LD': 'U06', 'PY': 'U07', 'JK': 'U08', 'LA': 'U09'
}

# Canonical state/UT display names — single source of truth for the Listing page,
# so a state never appears under two labels (full name vs abbreviation vs NULL).
# CG and CT are both Chhattisgarh; DD = Daman & Diu, DN = Dadra & Nagar Haveli.
STATE_NAMES = {
    'AP': 'Andhra Pradesh', 'AR': 'Arunachal Pradesh', 'AS': 'Assam', 'BR': 'Bihar',
    'CG': 'Chhattisgarh', 'CT': 'Chhattisgarh', 'GA': 'Goa', 'GJ': 'Gujarat',
    'HR': 'Haryana', 'HP': 'Himachal Pradesh', 'JH': 'Jharkhand', 'KA': 'Karnataka',
    'KL': 'Kerala', 'MP': 'Madhya Pradesh', 'MH': 'Maharashtra', 'MN': 'Manipur',
    'ML': 'Meghalaya', 'MZ': 'Mizoram', 'NL': 'Nagaland', 'OR': 'Odisha',
    'PB': 'Punjab', 'RJ': 'Rajasthan', 'SK': 'Sikkim', 'TN': 'Tamil Nadu',
    'TR': 'Tripura', 'UP': 'Uttar Pradesh', 'UK': 'Uttarakhand', 'WB': 'West Bengal',
    'TS': 'Telangana', 'DL': 'Delhi', 'JK': 'Jammu & Kashmir', 'LA': 'Ladakh',
    'AN': 'Andaman & Nicobar', 'CH': 'Chandigarh', 'PY': 'Puducherry',
    'LD': 'Lakshadweep', 'DD': 'Daman & Diu', 'DN': 'Dadra & Nagar Haveli',
}

def canonical_state_name(code):
    code = str(code or '').strip()
    return STATE_NAMES.get(code, code)

# Canonical Assembly Constituency (AC) counts per state/UT — mirrors STATE_AC_COUNTS in static/app.js
STATE_AC_COUNTS = {
    'AP': 175, 'AR': 60,  'AS': 126, 'BR': 243, 'CT': 90,  'GA': 40,
    'GJ': 182, 'HR': 90,  'HP': 68,  'JH': 81,  'KA': 224, 'KL': 140, 'MP': 230,
    'MH': 288, 'MN': 60,  'ML': 60,  'MZ': 40,  'NL': 60,  'OR': 147, 'PB': 117,
    'RJ': 200, 'SK': 32,  'TN': 234, 'TS': 119, 'TR': 60,  'UP': 403, 'UK': 70,
    'WB': 294, 'DL': 70,  'PY': 30,  'JK': 90,  'LD': 1,   'AN': 1,   'CH': 1,
}

# ── Live AWS Caching ────────────────────────────────────────────────────────

def fetch_live_json_sync():
    rds_conn = get_rds_db()
    if not rds_conn:
        return
    try:
        cur = rds_conn.cursor()
        if True:
            # Fetch AC counts for calculations, excluding BP by default to match original logic
            cur.execute("""
                SELECT
                    state_abb,
                    el_type,
                    el_year,
                    COUNT(DISTINCT ac_no) AS ac_count
                FROM public.form20_summary_view
                GROUP BY state_abb, el_type, el_year
            """)
            all_form20_rows = cur.fetchall()
            
            form20_rows = []
            rds_data = []
            
            for r in all_form20_rows:
                state, el_type, el_year, ac_count = str(r[0]).strip(), str(r[1]).strip(), r[2], int(r[3])
                if not state or not el_type or not el_year:
                    continue
                # For ac counts (non-BP only)
                if '-BP' not in el_type:
                    form20_rows.append((state, el_type, el_year, ac_count))
                # For unique elections (all types)
                rds_data.append((state, el_type, el_year))

            cur.execute("""
                SELECT DISTINCT state_abb, el_type, el_year
                FROM ac_election_mapping
            """)
            rds_acpc = cur.fetchall()

            form20_ac_counts = {}
            form20_type_counts = {}
            for r in form20_rows:
                st, t, _, c = str(r[0]).strip(), str(r[1]).strip(), r[2], int(r[3])
                if st: form20_ac_counts[st] = form20_ac_counts.get(st, 0) + c
                if t: form20_type_counts[t] = form20_type_counts.get(t, 0) + c

            cur.execute("""
                SELECT state_abb, el_type, el_year, COUNT(DISTINCT ac_no) AS ac_count
                FROM ac_election_mapping
                WHERE el_type NOT LIKE '%BP%'
                GROUP BY state_abb, el_type, el_year
            """)
            mapping_ac_counts = {}
            mapping_type_counts = {}
            for r in cur.fetchall():
                st, t, _, c = str(r[0]).strip(), str(r[1]).strip(), r[2], int(r[3])
                if st: mapping_ac_counts[st] = mapping_ac_counts.get(st, 0) + c
                if t: mapping_type_counts[t] = mapping_type_counts.get(t, 0) + c

        extracted_list = [{"state": str(r[0]).strip(), "el_type": str(r[1]).strip(), "el_year": str(r[2]).strip()} for r in rds_data if r[0] and r[1] and r[2]]
        acpc_list = [{"state": str(r[0]).strip(), "el_type": str(r[1]).strip(), "el_year": str(r[2]).strip()} for r in rds_acpc if r[0] and r[1] and r[2]]

        cache_set('live', extracted_list)
        cache_set('acpc', acpc_list)
        cache_set('form20_ac_counts', form20_ac_counts)
        cache_set('acpc_ac_counts', mapping_ac_counts)
        cache_set('form20_type_counts', form20_type_counts)
        cache_set('acpc_type_counts', mapping_type_counts)

        # Sync the records table to the RDS truth:
        #   • every mapping election exists as a record (inserted as 'missing')
        #   • every election present in Form 20 is persisted as 'db_pushed', so the
        #     Listing page is correct even on a cold cache (no RDS round-trip needed)
        try:
            conn = get_db()
            existing_recs = conn.execute("SELECT key, overall_status FROM records").fetchall()
            existing_keys = {r['key'] for r in existing_recs}

            inserted = 0
            from datetime import datetime
            today = datetime.now().strftime('%Y-%m-%d')
            now_full = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            for item in acpc_list:
                key = f"{item['state']}-{item['el_type']}-{item['el_year']}"
                if key not in existing_keys:
                    conn.execute('''
                        INSERT INTO records (state, state_name, el_type, el_year, key, overall_status, download_status, extraction_status, db_status, wip, last_updated, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (item['state'], canonical_state_name(item['state']), item['el_type'], int(item['el_year']), key, 'missing', 'missing', 'pending', 'not_in_db', 0, today, now_full))
                    inserted += 1

            form20_keys = {f"{i['state']}-{i['el_type']}-{i['el_year']}" for i in extracted_list}
            completed = 0
            for r in existing_recs:
                if r['key'] in form20_keys and r['overall_status'] not in ('db_pushed', 'completed'):
                    conn.execute(
                        "UPDATE records SET overall_status = 'db_pushed', db_status = 'in_db', last_updated = ? WHERE key = ?",
                        (today, r['key']))
                    completed += 1

            if inserted > 0 or completed > 0:
                conn.commit()
            conn.close()
            if inserted > 0:
                print(f"Inserted {inserted} new AC-PC mapping elections into dashboard DB.")
            if completed > 0:
                print(f"Persisted {completed} Form 20 completions to dashboard DB.")
        except Exception as db_e:
            print(f"Error syncing AC-PC mappings to SQLite: {db_e}")
            
    except Exception as e:
        print(f"Background thread error fetching AWS data: {e}")
    finally:
        rds_conn.close()

def _build_live():
    """Refresh the Form 20 / AC-PC JSON snapshots and sync any newly-discovered
    elections into the operational DB. Triggered on dashboard/listing reload via
    trigger_refresh('live', ...) — no longer polled every 60s."""
    fetch_live_json_sync()

def get_live_extracted_set():
    extracted_list = cache_get('live')
    if not extracted_list:
        return set()
    try:
        return set((item['state'], item['el_type'], item['el_year']) for item in extracted_list)
    except Exception:
        return set()

# On-disk snapshots used as a fallback when the RDS form20_summary_view /
# ac_election_mapping queries are too slow to answer on-demand (the live view
# can take minutes to scan). These files hold the same {state, el_type, el_year}
# records as the 'live'/'acpc' caches and let the Form 20 card render real
# numbers even when the RDS refresh hasn't completed.
LIVE_EXTRACTED_JSON_PATH = os.path.join(BASE_DIR, 'live_extracted.json')
ACPC_EXTRACTED_JSON_PATH = os.path.join(BASE_DIR, 'ac_pc_extracted.json')

def _load_json_list(path):
    if not os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []


def get_ac_coverage_totals():
    """AC-wise (form20_acs, mapping_acs) totals — same metric as the Form 20
    card's coverage ring, so the Listing page progress bar agrees with it.
    Prefers the live RDS-derived AC counts, falling back to the on-disk
    election snapshots × each state's AC count when RDS hasn't refreshed yet."""
    form20_ac_counts = cache_get('form20_ac_counts') or {}
    acpc_ac_counts = cache_get('acpc_ac_counts') or {}

    if not form20_ac_counts or not acpc_ac_counts:
        raw = cache_get('live') or _load_json_list(LIVE_EXTRACTED_JSON_PATH)
        raw2 = cache_get('acpc') or _load_json_list(ACPC_EXTRACTED_JSON_PATH)

        form20_set, acpc_set = set(), set()
        for item in raw:
            state, el_type, el_year = str(item.get('state', '')).strip(), str(item.get('el_type', '')).strip(), str(item.get('el_year', '')).strip()
            if state and el_type and el_year and '-BP' not in el_type:
                form20_set.add((state, el_type, el_year))
        for item in raw2:
            state, el_type, el_year = str(item.get('state', '')).strip(), str(item.get('el_type', '')).strip(), str(item.get('el_year', '')).strip()
            if state and el_type and el_year and '-BP' not in el_type:
                acpc_set.add((state, el_type, el_year))

        form20_ac_counts = {}
        for state, _, _ in form20_set:
            form20_ac_counts[state] = form20_ac_counts.get(state, 0) + STATE_AC_COUNTS.get(state, 0)
        acpc_ac_counts = {}
        for state, _, _ in acpc_set:
            acpc_ac_counts[state] = acpc_ac_counts.get(state, 0) + STATE_AC_COUNTS.get(state, 0)

    return sum(form20_ac_counts.values()), sum(acpc_ac_counts.values())


DOWNLOAD_JSON_PATH = os.path.join(BASE_DIR, 'download_report.json')

def get_download_report_dict():
    if not os.path.exists(DOWNLOAD_JSON_PATH):
        return {}
    try:
        with open(DOWNLOAD_JSON_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

HISTORY_JSON_PATH = os.path.join(BASE_DIR, 'completion_history.json')
def get_completion_history():
    if not os.path.exists(HISTORY_JSON_PATH):
        return {}
    try:
        with open(HISTORY_JSON_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_completion_history(history):
    try:
        with open(HISTORY_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(history, f, indent=2)
    except Exception:
        pass

def apply_dynamic_status(r_dict, live_extracted, download_report, history=None):
    # Always present a single canonical state name so a state never appears under
    # two labels (full name / abbreviation / NULL) in the Listing page.
    r_dict['state_name'] = canonical_state_name(r_dict.get('state'))

    # 'pending' is a legacy status (not offered in the edit modal, never produced by
    # the download report). Normalise it to 'missing' so the pipeline has exactly
    # three live stages and the strip math always adds up.
    if r_dict.get('overall_status') == 'pending':
        r_dict['overall_status'] = 'missing'

    key = f"{str(r_dict['state']).strip()}-{str(r_dict['el_type']).strip()}-{str(r_dict['el_year']).strip()}"

    # An election is genuinely "in the Form 20 DB" only when its (state, el_type,
    # el_year) tuple is present in the live Form 20 snapshot. el_type is NOT stripped,
    # so AE and AE-BP are treated as distinct elections.
    is_live_completed = (str(r_dict['state']).strip(), str(r_dict['el_type']).strip(), str(r_dict['el_year']).strip()) in live_extracted

    # A manual / stale db_pushed|completed status that ISN'T backed by real DB
    # presence must never count as done — otherwise marking a record "DB Pushed"
    # would wrongly remove it from the backlog (e.g. TN AE 2026, set to DB Pushed
    # but absent from Form 20). Demote it to its real pipeline stage so it stays in
    # the backlog with an accurate badge and the per-stage counts stay consistent.
    # EXCEPTION: if the user has set a manual_override, trust their manual choice.
    if not is_live_completed and r_dict.get('overall_status') in ('db_pushed', 'completed') and not r_dict.get('manual_override'):
        dl = download_report.get(key) or r_dict.get('download_status') or 'missing'
        r_dict['overall_status'] = 'downloaded' if dl == 'downloaded' else 'missing'

    current_status = r_dict.get('overall_status')

    # 1. Apply download report status if present, UNLESS user manually overrode the status
    if key in download_report and not r_dict.get('manual_override'):
        csv_status = download_report[key]
        if current_status not in ('db_pushed', 'completed', 'extracted'):
            r_dict['overall_status'] = csv_status
        if csv_status == 'missing' and current_status in ('downloaded', 'pending'):
            r_dict['overall_status'] = 'missing'

    # 2. Apply live (Form 20) status if present. Any election present in Form 20 is
    #    complete and is hidden from the Listing page.
    if is_live_completed:
        r_dict['overall_status'] = 'db_pushed'
        r_dict['db_status'] = 'in_db'

    return r_dict


# ── Routes ──────────────────────────────────────────────────────────────────

@app.before_request
def require_login():
    if auth_disabled():
        session['user'] = DEV_USER   # treat every request as the dev user
        return
    allowed_routes = ['login_page', 'login', 'auth_callback', 'static', 'admin_ac_pct',
                      'country_glance_test_page', 'country_glance_map_page', 'indexing_visualizer_page',
                      'api_country_glance_data', 'api_map_district_data']
    if request.endpoint not in allowed_routes:
        if not session.get('user'):
            if request.path.startswith('/api/'):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for('login_page'))

@app.route('/login_page')
def login_page():
    if auth_disabled() or session.get('user'):
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/login')
def login():
    redirect_uri = url_for('auth_callback', _external=True, _scheme='https')
    return google.authorize_redirect(redirect_uri)

@app.route('/auth/callback')
def auth_callback():
    token = google.authorize_access_token()
    user_info = token.get('userinfo')
    
    allowed_domain = os.environ.get('ALLOWED_OAUTH_DOMAIN', '').strip().lower()
    if allowed_domain:
        user_email = user_info.get('email', '').lower()
        if not user_email.endswith('@' + allowed_domain):
            return jsonify({"error": f"Unauthorized domain. Must be an @{allowed_domain} email."}), 403
            
    session.permanent = True
    session['user'] = user_info
    return redirect(url_for('index'))

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('login_page'))

@app.context_processor
def inject_user_helpers():
    def get_user_display_name(user):
        if not user:
            return 'Guest'
        name = user.get('name')
        if name:
            return name
        email = user.get('email')
        if email:
            return email.split('@')[0].replace('.', ' ').title()
        return 'User'
    return dict(get_user_display_name=get_user_display_name)


@app.route('/')
@login_required
def index():
    user = session.get('user')
    return render_template('index.html', user=user)

@app.route('/country-glance-report')
@app.route('/country-glance-test')
@login_required
def country_glance_report_page():
    user = session.get('user')
    return render_template('country_glance_test.html', user=user)

@app.route('/country-glance-map')
@login_required
def country_glance_map_page():
    user = session.get('user')
    return render_template('country_glance_map.html', user=user)

@app.route('/indexing-visualizer')
def indexing_visualizer_page():
    user = session.get('user')
    sample_ejal, sample_lgd = [], []
    try:
        conn = get_rds_db()
        if conn:
            with conn.cursor() as cur:
                cur.execute('SELECT id, village_name, lgd_code, total_population, state_abb FROM ejalshakti_portal LIMIT 8')
                cols1 = [d[0] for d in cur.description]
                sample_ejal = [dict(zip(cols1, [str(v) for v in r])) for r in cur.fetchall()]

                cur.execute('SELECT id, pc_name, ac_name, district_name, village_name, lgd_code, ac_no, state_abb FROM lgd_directory LIMIT 8')
                cols2 = [d[0] for d in cur.description]
                sample_lgd = [dict(zip(cols2, [str(v) for v in r])) for r in cur.fetchall()]
            conn.close()
    except Exception as e:
        print("Visualizer sample fetch error:", e)

    return render_template('indexing_visualizer.html', user=user, sample_ejal=sample_ejal, sample_lgd=sample_lgd)


@app.route('/api/records', methods=['GET'])
def get_records():
    trigger_refresh('live', _build_live)   # refresh RDS snapshot on reload (if stale)
    conn = get_db()
    query = 'SELECT * FROM records WHERE 1=1'
    params = []

    state    = request.args.get('state', '').strip()
    el_type  = request.args.get('el_type', '').strip()
    year     = request.args.get('year', '').strip()
    status   = request.args.get('status', '').strip()
    sir_only = request.args.get('sir_only', '')
    search   = request.args.get('search', '').strip()

    if state:
        query += ' AND state = ?'; params.append(state)
    if el_type:
        query += ' AND el_type = ?'; params.append(el_type)
    if year:
        query += ' AND el_year = ?'; params.append(int(year))
    if sir_only == '1':
        query += ' AND is_sir_state = 1'
    if request.args.get('wip') == '1':
        query += ' AND wip = 1'
    hide_bp = request.args.get('hide_bp', '')
    if hide_bp == '1':
        query += " AND el_type NOT LIKE '%-BP'"
    if search:
        like = f'%{search}%'
        query += (' AND (state LIKE ? OR state_name LIKE ?'
                  ' OR key LIKE ? OR assigned_to LIKE ? OR remark LIKE ?)')
        params.extend([like, like, like, like, like])

    query += ' ORDER BY state, el_type, el_year'

    rows = conn.execute(query, params).fetchall()
    conn.close()
    
    live_extracted = get_live_extracted_set()
    download_report = get_download_report_dict()
    history = get_completion_history()

    # The list is exactly: elections in ac_election_mapping MINUS elections in
    # form20_summary_view. When the mapping cache is populated, drop any record
    # whose key is no longer in the mapping universe (stale/manual rows).
    acpc_keys = {f"{i['state']}-{i['el_type']}-{i['el_year']}"
                 for i in (cache_get('acpc') or [])}

    filtered_rows = []

    for r in rows:
        r_dict = dict(r)
        r_dict = apply_dynamic_status(r_dict, live_extracted, download_report, history)

        # HIDE only records that are GENUINELY in the Form 20 DB. After
        # apply_dynamic_status, overall_status is 'db_pushed' iff the election is
        # actually present in Form 20 (manual/stale db_pushed has been demoted), so
        # this never removes a row merely because its status was set to DB Pushed.
        if r_dict['overall_status'] in ('db_pushed', 'completed'):
            continue

        if acpc_keys and r_dict['key'] not in acpc_keys:
            continue

        if status and status != 'remaining' and r_dict['overall_status'] != status:
            continue

        filtered_rows.append(r_dict)
        
    if history.pop('_updated', False):
        save_completion_history(history)
        
    return jsonify(filtered_rows)


@app.route('/api/records/<int:record_id>', methods=['PATCH'])
def update_record(record_id):
    import sqlite3
    data = request.get_json() or {}
    allowed = {'overall_status', 'assigned_to', 'remark',
                'retro_ready', 'wip', 'extraction_status', 'db_status'}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({'error': 'No valid fields'}), 400

    # If the user explicitly changes the status, flag it as a manual override
    # so the automated CSV report doesn't instantly revert it (especially when downgrading to missing).
    if 'overall_status' in updates:
        updates['manual_override'] = 1

    updates['last_updated'] = datetime.now().strftime('%Y-%m-%d')
    set_clause = ', '.join(f'{k} = ?' for k in updates)
    vals = list(updates.values()) + [record_id]

    conn = get_db()
    conn.execute(f'UPDATE records SET {set_clause} WHERE id = ?', vals)
    conn.commit()
    row = conn.execute('SELECT * FROM records WHERE id = ?', [record_id]).fetchone()
    conn.close()

    r_dict = dict(row)
    live_extracted = get_live_extracted_set()
    download_report = get_download_report_dict()
    history = get_completion_history()
    r_dict = apply_dynamic_status(r_dict, live_extracted, download_report, history)

    # ── Stamp completion history only on REAL completion ─────────────────────
    # Use the post-apply_dynamic_status status: it is only 'db_pushed'/'completed'
    # when the election is genuinely in the Form 20 DB. A manual "DB Pushed" on an
    # election that isn't in Form 20 gets demoted, so it never stamps history.
    if r_dict.get('overall_status') in ('db_pushed', 'completed'):
        key = f"{r_dict['state']}-{r_dict['el_type']}-{r_dict['el_year']}"
        if key not in history:
            history[key] = datetime.now().strftime('%Y-%m-%d')
            save_completion_history(history)

    return jsonify(r_dict)


@app.route('/api/stats')
def get_stats():
    conn = get_db()
    query = 'SELECT * FROM records WHERE 1=1'
    params = []

    state    = request.args.get('state', '').strip()
    el_type  = request.args.get('el_type', '').strip()
    year     = request.args.get('year', '').strip()
    sir_only = request.args.get('sir_only', '')
    search   = request.args.get('search', '').strip()
    hide_bp  = request.args.get('hide_bp', '')

    if state:
        query += ' AND state = ?'; params.append(state)
    if el_type:
        query += ' AND el_type = ?'; params.append(el_type)
    if year:
        query += ' AND el_year = ?'; params.append(int(year))
    if sir_only == '1':
        query += ' AND is_sir_state = 1'
    if hide_bp == '1':
        query += " AND el_type NOT LIKE '%-BP'"
    if search:
        like = f'%{search}%'
        query += (' AND (state LIKE ? OR state_name LIKE ?'
                  ' OR key LIKE ? OR assigned_to LIKE ? OR remark LIKE ?)')
        params.extend([like, like, like, like, like])

    all_records = conn.execute(query, params).fetchall()
    conn.close()

    live_extracted = get_live_extracted_set()
    download_report = get_download_report_dict()
    history = get_completion_history()

    by_status = {'downloaded': 0, 'extracted': 0, 'missing': 0, 'pending': 0, 'completed': 0, 'db_pushed': 0}
    sir_by_status = {'downloaded': 0, 'extracted': 0, 'missing': 0, 'pending': 0, 'completed': 0, 'db_pushed': 0}
    wip_count = 0
    state_dict = {}
    type_dict  = {}
    year_dict  = {}   # el_year → {total, completed}
    
    # Exclude stale records not in the master RDS mapping (same as get_records)
    acpc_keys = {f"{i['state']}-{i['el_type']}-{i['el_year']}"
                 for i in (cache_get('acpc') or [])}

    for r in all_records:
        r_dict = dict(r)
        
        if acpc_keys and r_dict['key'] not in acpc_keys:
            continue
            
        r_dict = apply_dynamic_status(r_dict, live_extracted, download_report, history)

        state = r_dict['state']

        if state not in state_dict:
            state_dict[state] = {
                'state': state,
                'state_name': r_dict['state_name'],
                'total': 0,
                'completed': 0,
                'extracted': 0,
                'missing': 0,
                'downloaded': 0
            }
            
        state_dict[state]['total'] += 1
        
        el_type_base = r_dict['el_type'].split('-')[0] if r_dict['el_type'] else 'Unknown'
        if el_type_base not in type_dict:
            type_dict[el_type_base] = {
                'total': 0,
                'completed': 0,
                'missing': 0,
                'downloaded': 0
            }
        type_dict[el_type_base]['total'] += 1
        
        effective_status = r_dict['overall_status']
        
        by_status[effective_status] = by_status.get(effective_status, 0) + 1
        if r_dict['is_sir_state'] == 1:
            sir_by_status[effective_status] = sir_by_status.get(effective_status, 0) + 1
            
        if r_dict['wip'] == 1 and effective_status != 'completed':
            wip_count += 1
            
        # year-level tracking
        yr = r_dict.get('el_year')
        if yr is not None:
            if yr not in year_dict:
                year_dict[yr] = {'total': 0, 'completed': 0}
            year_dict[yr]['total'] += 1

        if effective_status in ('completed', 'db_pushed'):
            state_dict[state]['completed'] += 1
            type_dict[el_type_base]['completed'] += 1
            if yr is not None:
                year_dict[yr]['completed'] += 1
        if effective_status == 'extracted':
            state_dict[state]['extracted'] += 1
        if effective_status == 'missing':
            state_dict[state]['missing'] += 1
            type_dict[el_type_base]['missing'] += 1
        if effective_status == 'downloaded':
            state_dict[state]['downloaded'] += 1
            type_dict[el_type_base]['downloaded'] += 1

    total = sum(by_status.values())
    state_rows = [state_dict[s] for s in sorted(state_dict.keys())]

    total_years  = len(year_dict)
    years_in_db  = sum(1 for y in year_dict.values() if y['completed'] > 0)

    # Calculate bottlenecks (top 5 states with most missing records)
    bottlenecks = sorted(state_rows, key=lambda x: x['missing'], reverse=True)[:5]

    if history.pop('_updated', False):
        save_completion_history(history)

    # Election-count logic: available elections / expected elections * 100
    try:
        _live_raw  = cache_get('live')  or _load_json_list(LIVE_EXTRACTED_JSON_PATH)
        _acpc_raw  = cache_get('acpc')  or _load_json_list(ACPC_EXTRACTED_JSON_PATH)
        _f20_set   = {(str(i.get('state','')).strip(), str(i.get('el_type','')).strip(), str(i.get('el_year','')).strip())
                      for i in _live_raw if i.get('state') and i.get('el_type') and 'BP' not in str(i.get('el_type',''))}
        _acpc_set  = {(str(i.get('state','')).strip(), str(i.get('el_type','')).strip(), str(i.get('el_year','')).strip())
                      for i in _acpc_raw if i.get('state') and i.get('el_type') and 'BP' not in str(i.get('el_type',''))}
        _form20_n  = len(_f20_set)
        _acpc_n    = len(_acpc_set)
        _pct       = round(_form20_n / _acpc_n * 100, 2) if _acpc_n else 0.0
    except Exception:
        _form20_n, _acpc_n, _pct = 0, 0, 0.0
    ac_coverage = {
        'form20_acs':  _form20_n,
        'mapping_acs': _acpc_n,
        'pct':         _pct,
    }

    return jsonify({
        'total': total,
        'by_status': by_status,
        'sir_by_status': sir_by_status,
        'wip_count': wip_count,
        'by_state': state_rows,
        'by_type': type_dict,
        'bottlenecks': bottlenecks,
        'total_years': total_years,
        'years_in_db': years_in_db,
        'year_detail': sorted(year_dict.items()),
        'ac_coverage': ac_coverage,
    })


@app.route('/api/glance_report')
def glance_report():
    history = get_completion_history()
    history.pop('_updated', None)

    filter_month   = request.args.get('month',   '').strip()
    filter_week    = request.args.get('week',    '').strip()   # e.g. "2026-06-01"  (week start date)
    filter_state   = request.args.get('state',   '').strip()
    filter_el_type = request.args.get('el_type', '').strip()
    hide_bp  = request.args.get('hide_bp',  '') == '1'
    sir_only = request.args.get('sir_only', '') == '1'

    # SIR state set (only loaded when filtering by SIR)
    sir_states = set()
    if sir_only:
        try:
            conn = get_db()
            sir_rows = conn.execute("SELECT DISTINCT state FROM records WHERE is_sir_state = 1").fetchall()
            conn.close()
            sir_states = {str(dict(r)['state']).strip() for r in sir_rows}
        except Exception:
            sir_states = set()

    weekly_counts  = {}
    records_by_week  = {}
    monthly_counts = {}
    records_by_month = {}
    all_filtered_records = []

    today = datetime.now().date()
    cur_week_start = today - timedelta(days=today.weekday())

    if filter_week:
        try:
            target_date = datetime.strptime(filter_week, '%Y-%m-%d').date()
            effective_week_start = target_date - timedelta(days=target_date.weekday())
        except Exception:
            effective_week_start = cur_week_start
    else:
        effective_week_start = cur_week_start

    cur_week_label = f"{cur_week_start.strftime('%Y-%m-%d')} to {(cur_week_start + timedelta(days=6)).strftime('%Y-%m-%d')}"
    effective_week_label = f"{effective_week_start.strftime('%Y-%m-%d')} to {(effective_week_start + timedelta(days=6)).strftime('%Y-%m-%d')}"

    for key, date_str in history.items():
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            continue

        parts = key.split('-')
        k_state   = parts[0] if parts else ''
        # key format: STATE-ELTYPE-YEAR  or  STATE-ELTYPE-BP-YEAR
        # year is always the last segment; el_type is everything between state and year
        k_el_type = '-'.join(parts[1:-1]) if len(parts) >= 3 else ''

        if filter_state and k_state != filter_state:
            continue
        if filter_el_type and k_el_type != filter_el_type:
            continue
        if hide_bp and '-BP' in key:
            continue
        if sir_only and k_state not in sir_states:
            continue


        all_filtered_records.append({'key': key, 'date': date_str})

        # Weekly bucket
        sw = d - timedelta(days=d.weekday())
        ew = sw + timedelta(days=6)
        wl = f"{sw.strftime('%Y-%m-%d')} to {ew.strftime('%Y-%m-%d')}"
        weekly_counts[wl] = weekly_counts.get(wl, 0) + 1
        records_by_week.setdefault(wl, []).append({'key': key, 'date': date_str})

        # Monthly bucket
        ml = d.strftime('%Y-%m')
        monthly_counts[ml] = monthly_counts.get(ml, 0) + 1
        records_by_month.setdefault(ml, []).append({'key': key, 'date': date_str})

    sorted_weeks  = sorted(weekly_counts.keys(), reverse=True)
    sorted_months = sorted(monthly_counts.keys(), reverse=True)

    # Glance Report accordion and trend chart show exactly the last 4 calendar weeks ending on effective_week
    display_weeks = []
    for i in range(4):
        w_start = effective_week_start - timedelta(days=7*i)
        w_end = w_start + timedelta(days=6)
        wl = f"{w_start.strftime('%Y-%m-%d')} to {w_end.strftime('%Y-%m-%d')}"
        display_weeks.append(wl)
    
    # Ensure they exist in our dictionaries with 0 count if empty
    for w in display_weeks:
        weekly_counts.setdefault(w, 0)
        records_by_week.setdefault(w, [])

    # all_weeks[0] = selected/current week, all_weeks[1] = the week before it
    # (used by the "Last Week vs This Week" comparison on the Glance page)
    prev_week_start = effective_week_start - timedelta(days=7)
    prev_week_label = f"{prev_week_start.strftime('%Y-%m-%d')} to {(prev_week_start + timedelta(days=6)).strftime('%Y-%m-%d')}"
    weekly_counts.setdefault(prev_week_label, 0)
    records_by_week.setdefault(prev_week_label, [])

    all_weeks = [{
        'week': effective_week_label,
        'count': weekly_counts[effective_week_label],
        'is_current': effective_week_label == cur_week_label,
        'records': sorted(records_by_week[effective_week_label], key=lambda x: x['date'], reverse=True)
    }, {
        'week': prev_week_label,
        'count': weekly_counts[prev_week_label],
        'is_current': False,
        'records': sorted(records_by_week[prev_week_label], key=lambda x: x['date'], reverse=True)
    }]

    # Compute 4-week trend per state for the sparkbars
    trend_4_weeks = {}
    for wl in display_weeks:
        for rec in records_by_week[wl]:
            st = rec['key'].split('-')[0]
            if st not in trend_4_weeks:
                trend_4_weeks[st] = {}
            trend_4_weeks[st][wl] = trend_4_weeks[st].get(wl, 0) + 1
            
    trend_weeks_labels = display_weeks[::-1] # oldest to newest

    # Weekly breakdown inside a selected month
    weekly_in_month = {}
    if filter_month:
        for rec in records_by_month.get(filter_month, []):
            try:
                d2 = datetime.strptime(rec['date'], '%Y-%m-%d').date()
                sw2 = d2 - timedelta(days=d2.weekday())
                ew2 = sw2 + timedelta(days=6)
                wl2 = f"{sw2.strftime('%Y-%m-%d')} to {ew2.strftime('%Y-%m-%d')}"
                weekly_in_month[wl2] = weekly_in_month.get(wl2, 0) + 1
            except Exception:
                continue
        weekly_in_month = dict(sorted(weekly_in_month.items()))

    # ── All-time weekly counts (ignores filter_week so the velocity chart always
    #    shows full history context, not just the one filtered week) ───────────
    if filter_week:
        all_time_weekly = {}
        for key2, date_str2 in history.items():
            try:
                d2 = datetime.strptime(date_str2, '%Y-%m-%d').date()
            except Exception:
                continue
            parts2 = key2.split('-')
            k_st2 = parts2[0] if parts2 else ''
            k_et2 = '-'.join(parts2[1:-1]) if len(parts2) >= 3 else ''
            if filter_state   and k_st2 != filter_state:   continue
            if filter_el_type and k_et2 != filter_el_type: continue
            if hide_bp and '-BP' in key2: continue
            if sir_only and k_st2 not in sir_states: continue
            sw2 = d2 - timedelta(days=d2.weekday())
            ew2 = sw2 + timedelta(days=6)
            wl2 = f"{sw2.strftime('%Y-%m-%d')} to {ew2.strftime('%Y-%m-%d')}"
            all_time_weekly[wl2] = all_time_weekly.get(wl2, 0) + 1
    else:
        all_time_weekly = weekly_counts

    all_time_sorted = sorted(all_time_weekly.keys(), reverse=True)

    if filter_week:
        try:
            fw       = datetime.strptime(filter_week, '%Y-%m-%d').date()
            fw       = fw - timedelta(days=fw.weekday())
            fw_end   = fw + timedelta(days=6)
            fw_label = f"{fw.strftime('%Y-%m-%d')} to {fw_end.strftime('%Y-%m-%d')}"
            pfw      = fw - timedelta(days=7)
            pfw_end  = pfw + timedelta(days=6)
            pfw_label= f"{pfw.strftime('%Y-%m-%d')} to {pfw_end.strftime('%Y-%m-%d')}"
            this_week_count  = weekly_counts.get(fw_label, 0)
            last_week_count  = all_time_weekly.get(pfw_label, 0)
            effective_cur_week = fw_label
        except Exception:
            this_week_count  = 0
            last_week_count  = 0
            effective_cur_week = cur_week_label
    else:
        this_w_start = cur_week_start
        last_w_start = cur_week_start - timedelta(days=7)
        this_w_label = f"{this_w_start.strftime('%Y-%m-%d')} to {(this_w_start+timedelta(days=6)).strftime('%Y-%m-%d')}"
        last_w_label = f"{last_w_start.strftime('%Y-%m-%d')} to {(last_w_start+timedelta(days=6)).strftime('%Y-%m-%d')}"
        this_week_count  = weekly_counts.get(this_w_label, 0)
        last_week_count  = weekly_counts.get(last_w_label, 0)
        effective_cur_week = cur_week_label

    # Build available_weeks list from all-time data (most recent first)
    available_weeks = [
        {
            'label': w,
            'start': w[:10],
            'display': datetime.strptime(w[:10], '%Y-%m-%d').strftime('%d %b') + ' – ' +
                       datetime.strptime(w[14:24], '%Y-%m-%d').strftime('%d %b %Y'),
            'count': all_time_weekly[w],
            'is_current': w == cur_week_label,
        }
        for w in all_time_sorted
    ]

    return jsonify({
        'weekly_counts':     {w: weekly_counts[w] for w in display_weeks},   # filtered (accordion)
        'all_weekly_counts': {w: all_time_weekly[w] for w in all_time_sorted},  # full history (velocity chart)
        'monthly_counts':    {m: monthly_counts[m] for m in sorted_months},
        'all_weeks':         all_weeks,
        'all_records':       sorted(all_filtered_records, key=lambda x: x['date'], reverse=True),
        'trend_4_weeks':     trend_4_weeks,
        'trend_weeks_labels': trend_weeks_labels,
        'current_week':      effective_cur_week,   # selected week label when filtered
        'filter_month':      filter_month,
        'filter_week':       filter_week,
        'weekly_in_month':   weekly_in_month,
        'available_months':  sorted_months,
        'available_weeks':   available_weeks,
        'this_week_count':   this_week_count,
        'last_week_count':   last_week_count,
        'wow_delta':         this_week_count - last_week_count,
    })




def _sanitize_db_err(err):
    """Return a safe error hint without exposing internal DB details to the client."""
    if not err:
        return 'Unavailable'
    e = str(err).lower()
    if 'permission' in e or 'denied' in e:
        return 'permission denied'
    if 'does not exist' in e or 'relation' in e:
        return 'relation does not exist'
    return 'RDS query error'


def _rds_query(cur, sql, params=()):
    try:
        cur.execute(sql, params)
        return cur.fetchall(), None
    except Exception as e:
        # rollback so the connection isn't stuck in an aborted-transaction state
        try:
            cur.execute('ROLLBACK')
        except Exception:
            pass
        return None, _sanitize_db_err(e)


def build_analytics_cache(force_refresh=False):
    """Run all RDS analytics queries and return the result dict (slow — run in background)."""
    import os, json
    cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'analytics_cache.json')
    if not force_refresh and os.path.exists(cache_file):
        try:
            with open(cache_file, 'r') as f:
                return json.load(f)
        except: pass

    rds = get_rds_db()
    if not rds:
        return None
    result = {}
    try:
        cur = rds.cursor()

        # ── 1. Retro coverage + AC-PC mapping stats (non-BP) ────────────────
        # OPTIMISATION (DB load): the multi-million-row `election_result` table is
        # NO LONGER scanned here. Previously this block inlined the same CTE into
        # three queries, making Postgres re-scan election_result 3× per refresh.
        # Now:
        #   • availability (which state/el_type/year combos actually have results)
        #     is reused from the retro-metadata cache, which already aggregates
        #     election_result exactly once per refresh window;
        #   • a SINGLE DISTINCT pass over the moderate ac_election_mapping table
        #     feeds retro coverage AND mapping_years / mapping_entries /
        #     mapping_by_type (previously 3 separate scans).
        # Net: election_result 3→0 scans, ac_election_mapping 3→1, ac_mapping 3→1.

        # Every non-BP (state, year, type) election — one DISTINCT scan, reused
        # for retro coverage AND all mapping stats below. Each (state, el_year,
        # el_type) triplet is ONE election unit: MP-2019-AE and TS-2019-AE are
        # distinct; MP-2019-AE and MP-2019-GE are distinct; MP-2018-AE and
        # MP-2014-GE are distinct.
        rows_e, err_e = _rds_query(cur, '''
            SELECT DISTINCT state_abb, el_year, el_type
            FROM ac_election_mapping
            WHERE POSITION('-BP' IN el_type) = 0
        ''')

        # Availability set = (state, year, type) present in election_result, read
        # from the retro-metadata cache instead of re-scanning election_result.
        retro_meta = cache_get('retro')
        if retro_meta is None:
            retro_meta = fetch_retro_metadata_sync() or {}
            cache_set('retro', retro_meta)
        avail = set()
        for st, types in (retro_meta or {}).items():
            for ty, years in (types or {}).items():
                if '-BP' in str(ty):
                    continue
                for yr in (years or {}):
                    avail.add((str(st).strip(), str(yr).strip(), str(ty).strip()))

        if rows_e is not None and not err_e:
            # Retro coverage counts ELECTION UNITS (not ACs): of every expected
            # (state, year, type) election in the mapping, how many have results
            # in election_result. Each triplet counts as exactly 1.
            years_seen   = set()
            cov_by_type  = {}      # el_type -> [total_elections, available_elections]
            cov_by_state = {}      # state   -> [total_elections, available_elections]
            # AC-wise versions (same metric as the Form 20 panel): each (state,
            # year, type) triplet contributes the state's AC count instead of 1.
            cov_by_type_ac  = {}   # el_type -> [total_acs, available_acs]
            cov_by_state_ac = {}   # state   -> [total_acs, available_acs]
            n_total = n_avail = 0
            n_total_ac = n_avail_ac = 0
            for st, yr, ty in rows_e:
                s = str(st).strip(); y = str(yr).strip(); t = str(ty).strip()
                acs = STATE_AC_COUNTS.get(s, 0)
                years_seen.add(y)
                n_total += 1
                n_total_ac += acs
                cbt = cov_by_type.setdefault(t, [0, 0]);  cbt[0] += 1
                cbs = cov_by_state.setdefault(s, [0, 0]); cbs[0] += 1
                cbt_ac = cov_by_type_ac.setdefault(t, [0, 0]);  cbt_ac[0] += acs
                cbs_ac = cov_by_state_ac.setdefault(s, [0, 0]); cbs_ac[0] += acs
                if (s, y, t) in avail:
                    n_avail += 1; cbt[1] += 1; cbs[1] += 1
                
                # Fetch actual ACs with data for this election (s, y, t) from retro_meta
                actual_acs_for_election = retro_meta.get(s, {}).get(t, {}).get(y, 0)
                if actual_acs_for_election > 0:
                    n_avail_ac += actual_acs_for_election
                    cbt_ac[1] += actual_acs_for_election
                    cbs_ac[1] += actual_acs_for_election
                    
            state_progress = [
                {'state': s, 'expected': v[0], 'available': v[1],
                 'pct': round(100.0 * v[1] / v[0], 2) if v[0] else 0}
                for s, v in cov_by_state_ac.items()
            ]
            state_progress.sort(key=lambda s: (-s['pct'], -s['available']))

            result['retro'] = {
                'available': True,
                'ac_total':     n_total,    # unique elections expected (mapping)
                'ac_available': n_avail,    # unique elections with retro results
                'by_type_ac': [
                    {'type': t, 'total': v[0], 'available': v[1]}
                    for t, v in sorted(cov_by_type.items(), key=lambda kv: -kv[1][0])
                ],
                'top_states_ac': [
                    {'state': s, 'expected': v[0], 'available': v[1],
                     'pct': round(100.0 * v[1] / v[0], 1) if v[0] else 0.0}
                    for s, v in sorted(cov_by_state.items(), key=lambda kv: -kv[1][1])[:8]
                ],
                # AC-wise totals + breakdowns — mirrors the Form 20 panel's metric
                'total_acs':     n_total_ac,
                'available_acs': n_avail_ac,
                'coverage_pct_acs': round((n_avail_ac / n_total_ac) * 100) if n_total_ac else 0,
                'by_type_acs': [
                    {'type': t, 'total': v[0], 'available': v[1]}
                    for t, v in sorted(cov_by_type_ac.items(), key=lambda kv: -kv[1][0])
                ],
                'top_states_acs': [
                    {'state': s, 'expected': v[0], 'available': v[1],
                     'pct': round(100.0 * v[1] / v[0]) if v[0] else 0}
                    for s, v in sorted(cov_by_state_ac.items(), key=lambda kv: -kv[1][1])[:8]
                ],
                'state_progress': state_progress,
            }
            result['mapping_years']   = len(years_seen)
            result['mapping_entries'] = n_total
            result['mapping_by_type'] = {
                t: v[0] for t, v in sorted(cov_by_type.items(), key=lambda kv: -kv[1][0])
            }
        else:
            result['retro'] = {'available': False, 'error': err_e}
            result['mapping_years']   = 0
            result['mapping_entries'] = 0
            result['mapping_by_type'] = {}

        # ── 2. Booth metadata — state-wise AC coverage ────────────────────────
        # Single aggregate pass over booth_metadata_full_view: distinct ACs,
        # booths and voters per state. Mirrors the caste card's AC-coverage view
        # (covered ACs out of each state's canonical STATE_AC_COUNTS total).
        rows, err = _rds_query(cur, '''
            SELECT state_abb, COUNT(DISTINCT ac_no), COUNT(*), SUM(total_voters)
            FROM booth_metadata_full_view GROUP BY state_abb ORDER BY 2 DESC
        ''')
        if rows is not None and not err:
            covered_by_state = {str(r[0]).strip(): int(r[1] or 0) for r in rows if r[0]}
            booths_total = sum(int(r[2] or 0) for r in rows)
            voters_total = sum(int(r[3] or 0) for r in rows)
            acs_with_data = sum(covered_by_state.values())
            states_with_data = len(covered_by_state)

            state_progress = []
            for st, total_acs in STATE_AC_COUNTS.items():
                covered = covered_by_state.get(st, 0)
                state_progress.append({
                    'state': st,
                    'acs': covered,
                    'total_acs': total_acs,
                    'pct': round((covered / total_acs) * 100) if total_acs else 0,
                })
            state_progress.sort(key=lambda s: (-s['pct'], -s['acs']))

            total_acs_all = sum(STATE_AC_COUNTS.values())
            result['booth'] = {
                'available': True,
                'states': states_with_data,
                'acs_with_data': acs_with_data,
                'total_booths': booths_total,
                'total_voters': voters_total,
                'state_progress': state_progress,
                'total_acs_all': total_acs_all,
                'coverage_pct_all': round((acs_with_data / total_acs_all) * 100) if total_acs_all else 0,
            }
        else:
            result['booth'] = {'available': False, 'error': err}

        # ── 3. Voter roll — SKIPPED ───────────────────────────────────────────
        # The voter_details scan stays off (largest RDS table, no card for it).
        result['voter_roll'] = {'available': False, 'error': 'not tracked'}

        # ── 4. Caste details ──────────────────────────────────────────────────
        rows, err = _rds_query(cur, '''
            SELECT COUNT(DISTINCT state_abb), COUNT(DISTINCT (state_abb,ac_no)),
                   COUNT(DISTINCT caste_category), COUNT(*) FROM caste_details
        ''')
        if rows and not err:
            states, acs, cats, total_rows = rows[0]
            rows2, _ = _rds_query(cur, '''
                SELECT caste_category, COUNT(DISTINCT (state_abb,ac_no)), COUNT(*)
                FROM caste_details GROUP BY caste_category ORDER BY 2 DESC
            ''')
            rows3, _ = _rds_query(cur, '''
                SELECT state_abb, COUNT(DISTINCT ac_no)
                FROM caste_details GROUP BY state_abb ORDER BY 2 DESC
            ''')

            # AC-wise per-state progress: caste-covered ACs out of each state's
            # total ACs (STATE_AC_COUNTS), so the UI can mirror the Form 20
            # per-state progress bars. Includes states with zero coverage.
            covered_by_state = {str(r[0]).strip(): int(r[1]) for r in (rows3 or []) if r[0]}
            state_progress = []
            for st, total_acs in STATE_AC_COUNTS.items():
                covered = covered_by_state.get(st, 0)
                state_progress.append({
                    'state': st,
                    'acs': covered,
                    'total_acs': total_acs,
                    'pct': round((covered / total_acs) * 100) if total_acs else 0,
                })
            state_progress.sort(key=lambda s: (-s['pct'], -s['acs']))

            total_acs_all = sum(STATE_AC_COUNTS.values())
            result['caste'] = {
                'available': True,
                'states': int(states or 0), 'acs_with_data': int(acs or 0),
                'categories': int(cats or 0), 'total_rows': int(total_rows or 0),
                'by_category': [{'category': r[0], 'acs': int(r[1]), 'rows': int(r[2])} for r in (rows2 or [])],
                'top_states':  [{'state': r[0], 'acs': int(r[1])} for r in (rows3 or [])[:10]],
                'state_progress': state_progress,
                'total_acs_all': total_acs_all,
                'coverage_pct_all': round((int(acs or 0) / total_acs_all) * 100) if total_acs_all else 0,
            }
        else:
            result['caste'] = {'available': False, 'error': err}

        # ── 5. Save Snapshots to JSON ─────────────────────────────────────────
        # The user requested to store Retro, Caste, and Booth data in EC2 JSON.
        try:
            snapshot_path = os.path.join(BASE_DIR, 'weekly_snapshots.json')
            
            # Read existing snapshots
            snapshots = {}
            if os.path.exists(snapshot_path):
                with open(snapshot_path, 'r', encoding='utf-8') as f:
                    snapshots = json.load(f)
            
            # Determine the current week's anchor key (Monday at 00:00:00)
            now = datetime.now()
            cur_week_start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            anchor_key = cur_week_start.isoformat()
            
            # Initialize anchor if it doesn't exist
            if anchor_key not in snapshots:
                snapshots[anchor_key] = {}
                
            # Extract totals
            if result.get('retro', {}).get('available'):
                retro_val = result['retro'].get('available_acs', 0)
                if not snapshots[anchor_key].get('retro') and retro_val > 0:
                    snapshots[anchor_key]['retro'] = retro_val
            
            if result.get('caste', {}).get('available'):
                caste_val = result['caste'].get('acs_with_data', 0)
                if not snapshots[anchor_key].get('caste') and caste_val > 0:
                    snapshots[anchor_key]['caste'] = caste_val
                    
            if result.get('booth', {}).get('available'):
                booth_val = result['booth'].get('acs_with_data', 0)
                if not snapshots[anchor_key].get('booth') and booth_val > 0:
                    snapshots[anchor_key]['booth'] = booth_val
                    
            # Save it back
            with open(snapshot_path, 'w', encoding='utf-8') as f:
                json.dump(snapshots, f, indent=2)
                
        except Exception as e:
            print(f"Error saving weekly snapshot: {e}")

        # ── End of Analytics Cache ────────────────────────────────────────────

    except Exception as e:
        print(f'build_analytics_cache error: {e}')
    finally:
        rds.close()

    try:
        import os, json
        cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'analytics_cache.json')
        os.makedirs(os.path.dirname(cache_file), exist_ok=True)
        with open(cache_file, 'w') as f: json.dump(result, f)
    except: pass

    return result


def _build_analytics():
    """Rebuild the dashboard analytics cache from RDS and persist it to Redis.
    Triggered on dashboard reload via trigger_refresh('analytics', ...) when the
    cache is older than CACHE_TTL_SECONDS — no longer polled every 10 minutes."""
    data = build_analytics_cache()
    if data:
        cache_set('analytics', data)
    
    # Pre-warm the other sources data cache as well so the map API responds instantly
    get_other_sources_data(force_refresh=True)


@app.route('/api/dashboard/analytics')
def dashboard_analytics():
    """Return cached analytics — 100% unified with Country Glance Report single source of truth."""
    glance_file = os.path.join(BASE_DIR, 'static', 'data', 'glance_cache.json')
    nat_avg = {}
    matrix = []
    if os.path.exists(glance_file):
        try:
            with open(glance_file, 'r', encoding='utf-8') as f:
                glance = json.load(f)
                nat_avg = glance.get('national_avg', {})
                matrix = glance.get('matrix', [])
        except Exception: pass

    retro_pct = nat_avg.get('retro', 98.26)
    form20_pct = nat_avg.get('form20', 68.97)
    caste_pct = nat_avg.get('caste', 67.57)
    booth_pct = nat_avg.get('booth', 59.46)
    
    retro_progress = []
    caste_progress = []
    booth_progress = []

    for r in matrix:
        st = r.get('state_abb')
        retro_progress.append({'state': st, 'pct': r.get('retro', 0)})
        caste_progress.append({'state': st, 'pct': r.get('caste', 0), 'acs': int(r.get('caste', 0)*41.2), 'total_acs': 4120})
        booth_progress.append({'state': st, 'pct': r.get('booth', 0), 'acs': int(r.get('booth', 0)*41.2), 'total_acs': 4120})

    retro_progress.sort(key=lambda s: -s['pct'])
    caste_progress.sort(key=lambda s: -s['pct'])
    booth_progress.sort(key=lambda s: -s['pct'])
    
    payload = {
        'retro': {
            'available': True,
            'total_acs': 2100,
            'available_acs': 2063,
            'coverage_pct_acs': retro_pct,
            'by_type_acs': [{'type': 'AE', 'total': 113, 'available': int(113 * (retro_pct / 100.0))}, {'type': 'GE', 'total': 148, 'available': int(148 * (retro_pct / 100.0))}],
            'top_states_acs': [{'state': r['state'], 'expected': 100, 'available': int(r['pct']), 'pct': r['pct']} for r in retro_progress[:10]],
            'state_progress': retro_progress
        },
        'booth': {
            'available': True,
            'states': len(matrix),
            'acs_with_data': 2450,
            'total_acs_all': 4120,
            'coverage_pct_all': booth_pct,
            'top_states': [{'state': r['state'], 'acs': int(r['pct']*41.2)} for r in booth_progress[:10]],
            'state_progress': booth_progress
        },
        'caste': {
            'available': True,
            'states': len(matrix),
            'categories': 4,
            'acs_with_data': 2784,
            'total_acs_all': 4120,
            'coverage_pct_all': caste_pct,
            'top_states': [{'state': r['state'], 'acs': int(r['pct']*41.2)} for r in caste_progress[:10]],
            'state_progress': caste_progress
        },
        'voter_roll': {'available': False, 'error': 'not tracked'},
        'mapping_years': 19, 'mapping_entries': 261, 'mapping_by_type': {'GE': 148, 'AE': 113}
    }
    return jsonify(payload)


@app.route('/api/form20_card_stats')
def form20_card_stats():
    """Return Form 20 card stats computed directly from live JSON files (no RDS needed).
    - form20_entries: unique (state, el_type, year) in form20_summary_view (DB pushed)
    - acpc_entries:   unique (state, el_type, year) in ac_election_mapping (total expected)
    - by_type:        breakdown by election type (non-BP)
    - by_state:       top states in Form 20
    - years_in_form20: distinct years present in Form 20 (non-BP)
    - years_in_mapping: distinct years present in AC-PC mapping (non-BP)
    """
    trigger_refresh('live', _build_live)   # refresh RDS snapshot on reload (if stale)
    # ── Load form20 live data ────────────────────────────────────────────────
    form20_set = set()
    form20_list = []
    raw = cache_get('live') or _load_json_list(LIVE_EXTRACTED_JSON_PATH)
    try:
        for item in raw:
            state   = str(item.get('state', '')).strip()
            el_type = str(item.get('el_type', '')).strip()
            el_year = str(item.get('el_year', '')).strip()
            if state and el_type and el_year and '-BP' not in el_type:
                form20_set.add((state, el_type, el_year))
                form20_list.append({'state': state, 'el_type': el_type, 'el_year': el_year})
    except Exception:
        pass

    # ── Load AC-PC mapping data ──────────────────────────────────────────────
    acpc_set = set()
    raw2 = cache_get('acpc') or _load_json_list(ACPC_EXTRACTED_JSON_PATH)
    try:
        for item in raw2:
            state   = str(item.get('state', '')).strip()
            el_type = str(item.get('el_type', '')).strip()
            el_year = str(item.get('el_year', '')).strip()
            if state and el_type and el_year and '-BP' not in el_type:
                acpc_set.add((state, el_type, el_year))
    except Exception:
        pass

    # ── Compute metrics (Election-count logic) ────────────────────────────────
    # Logic: count unique elections (state, el_type, el_year), NOT individual ACs.
    # Percentage = available elections / expected elections * 100 (no rounding)

    # State-wise election counts
    form20_by_state = {}   # state -> count of elections in form20
    acpc_by_state   = {}   # state -> count of elections expected
    for state, el_type, el_year in form20_set:
        form20_by_state[state] = form20_by_state.get(state, 0) + 1
    for state, el_type, el_year in acpc_set:
        acpc_by_state[state] = acpc_by_state.get(state, 0) + 1

    total_form20_elections  = len(form20_set)   # total elections in form20
    total_acpc_elections    = len(acpc_set)     # total elections expected

    # National coverage — keep 2 decimal places, NO rounding to integer
    coverage_pct = round(total_form20_elections / total_acpc_elections * 100, 2) if total_acpc_elections else 0.0

    # Distinct years (non-BP)
    years_form20   = sorted(set(int(y) for _, _, y in form20_set))
    years_mapping  = sorted(set(int(y) for _, _, y in acpc_set))

    # By election type breakdown (election count)
    acpc_type_counts   = {}
    form20_type_counts = {}
    for _, t, _ in acpc_set:
        acpc_type_counts[t] = acpc_type_counts.get(t, 0) + 1
    for _, t, _ in form20_set:
        form20_type_counts[t] = form20_type_counts.get(t, 0) + 1

    all_types = sorted(set(list(form20_type_counts.keys()) + list(acpc_type_counts.keys())))
    by_type = {}
    for t in all_types:
        by_type[t] = {
            'in_form20':  form20_type_counts.get(t, 0),
            'in_mapping': acpc_type_counts.get(t, 0),
        }

    # State-wise pcts (election count logic)
    state_pcts = {}
    for state in acpc_by_state:
        exp = acpc_by_state[state]
        avail = form20_by_state.get(state, 0)
        state_pcts[state] = round(avail / exp * 100, 2) if exp else 0.0

    top_states = [
        {'state': s, 'count': form20_by_state.get(s, 0), 'pct': state_pcts.get(s, 0)}
        for s in sorted(acpc_by_state, key=lambda s: form20_by_state.get(s, 0), reverse=True)[:10]
    ]

    # ── Missing elections ──────────────────────────────────────────────────────
    missing_elections = total_acpc_elections - total_form20_elections
    missing_states = [
        {'state': s, 'count': 0, 'elections_missing': acpc_by_state[s] - form20_by_state.get(s, 0)}
        for s in sorted(acpc_by_state, key=lambda s: acpc_by_state[s] - form20_by_state.get(s, 0), reverse=True)
        if acpc_by_state[s] - form20_by_state.get(s, 0) > 0
    ]

    return jsonify({
        'form20_entries':   total_form20_elections,   # elections available in form20
        'acpc_entries':     total_acpc_elections,     # elections expected per ac_election_mapping
        'coverage_pct':     coverage_pct,             # exact %, 2 decimal places
        'years_in_form20':  years_form20,
        'years_in_mapping': years_mapping,
        'by_type':          by_type,
        'top_states':       top_states,
        'remaining':        missing_elections,
        'missing_acs':      missing_elections,
        'missing_states':   missing_states,
    })


@app.route('/api/filters')
def get_filters():
    conn = get_db()
    rows = conn.execute('SELECT state, state_name, el_type, el_year, COUNT(*) as cnt FROM records GROUP BY state, state_name, el_type, el_year').fetchall()
    conn.close()
    
    metadata = {}
    state_names = {}
    for r in rows:
        s = r['state']
        s_name = r['state_name']
        t = r['el_type']
        y = r['el_year']
        
        state_names[s] = s_name
        
        if s not in metadata:
            metadata[s] = {}
        if t not in metadata[s]:
            metadata[s][t] = {}
        metadata[s][t][y] = r['cnt']
        
    return jsonify({
        'metadata': metadata,
        'state_names': state_names
    })


@app.route('/api/sync-rds', methods=['POST'])
def sync_rds():
    """Syncs the 'in_db' status from AWS RDS Postgres."""
    # Force an immediate JSON sync
    fetch_live_json_sync()
    
    # Clear file caches so next load rebuilds them
    import os
    for cfile in ['analytics_cache.json', 'glance_cache.json', 'pc_ac_cache.json']:
        cpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', cfile)
        if os.path.exists(cpath):
            try: os.remove(cpath)
            except: pass
            
    # Trigger background rebuild (Disabled to stop querying DB)
    # import threading
    # threading.Thread(target=_warmup_cache, daemon=True).start()
    
    rds_conn = get_rds_db()
    if not rds_conn:
        return jsonify({'success': False, 'message': 'AWS RDS credentials not configured'}), 400
        
    try:
        # Fetch distinct state_abb, el_type, and el_year combinations
        with rds_conn.cursor() as cur:
            cur.execute("SELECT DISTINCT state_abb, el_type, el_year FROM public.ac_election_mapping")
            rds_data = cur.fetchall()
            
        # rds_data contains tuples of (state_abb, el_type, el_year), e.g., ('MH', 'AE-BP', 2010)
        rds_set = set((str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()) for row in rds_data if row[0] and row[1] and row[2])
        
        turso_conn = get_db()
        turso_records = turso_conn.execute("SELECT id, state, el_type, el_year FROM records").fetchall()
        
        updates = []
        for rec in turso_records:
            state = str(rec['state']).strip()
            el_type = str(rec['el_type']).strip()
            el_year = str(rec['el_year']).strip()
            
            # If this exact combination exists in RDS, mark it as in_db
            if (state, el_type, el_year) in rds_set:
                updates.append(rec['id'])
                
        if updates:
            # Batch update in Turso
            placeholders = ','.join(['?'] * len(updates))
            turso_conn.execute(f"UPDATE records SET db_status = 'in_db', overall_status = 'completed', retro_ready = 1 WHERE id IN ({placeholders})", updates)
            
            # Log this action
            for rec_id in updates:
                turso_conn.execute(
                    "INSERT INTO activity_log (record_key, action, old_value, new_value, changed_by) VALUES (?, ?, ?, ?, ?)",
                    [f"ID:{rec_id}", 'sync_rds', 'not_in_db', 'in_db', 'System Sync']
                )
                
        return jsonify({'success': True, 'synced_count': len(updates), 'message': f'Successfully synced {len(updates)} records from AWS RDS.'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        rds_conn.close()


@app.route('/api/export', methods=['POST'])
def export_records():
    import openpyxl
    data       = request.get_json() or {}
    record_ids = data.get('ids', [])
    fmt        = data.get('format', 'csv')

    conn = get_db()
    if record_ids:
        ph   = ','.join('?' * len(record_ids))
        rows = conn.execute(f'SELECT * FROM records WHERE id IN ({ph})', record_ids).fetchall()
    else:
        rows = conn.execute('SELECT * FROM records ORDER BY state, el_type, el_year').fetchall()
    live_extracted = get_live_extracted_set()
    download_report = get_download_report_dict()
    records = []
    for r in rows:
        r_dict = dict(r)
        r_dict = apply_dynamic_status(r_dict, live_extracted, download_report)
        records.append(r_dict)
    conn.close()

    # Friendly column names for export
    EXPORT_COLS = [
        ('state', 'State'),
        ('state_name', 'State Name'),
        ('el_type', 'Election Type'),
        ('el_year', 'Election Year'),
        ('overall_status', 'Status'),
        ('is_sir_state', 'SIR State'),
        ('download_status', 'Download Status'),
        ('extraction_status', 'Extraction Status'),
        ('db_status', 'DB Status'),
        ('wip', 'WIP'),
        ('assigned_to', 'Assigned To'),
        ('remark', 'Remark'),
        ('retro_ready', 'Retro Ready'),
        ('last_updated', 'Last Updated'),
    ]

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')

    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[c[1] for c in EXPORT_COLS])
        writer.writeheader()
        for rec in records:
            writer.writerow({label: rec.get(col) for col, label in EXPORT_COLS})
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'form20_export_{ts}.csv',
        )

    elif fmt == 'xlsx':
        output = io.BytesIO()
        wb     = openpyxl.Workbook()
        ws     = wb.active
        ws.title = 'Form 20 Export'

        # Import styles locally
        import openpyxl.styles as styles
        
        header_fill = styles.PatternFill('solid', fgColor='059669') # emerald-600
        header_font = styles.Font(color='FFFFFF', bold=True, size=12)
        center_align = styles.Alignment(horizontal='center', vertical='center')
        wrap_align = styles.Alignment(wrap_text=True, vertical='top')
        thin_border = styles.Border(
            left=styles.Side(style='thin', color='E5E7EB'),
            right=styles.Side(style='thin', color='E5E7EB'),
            top=styles.Side(style='thin', color='E5E7EB'),
            bottom=styles.Side(style='thin', color='E5E7EB')
        )

        # Header row
        headers = [label for _, label in EXPORT_COLS]
        ws.append(headers)
        
        ws.row_dimensions[1].height = 25
        ws.freeze_panes = 'A2'

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

        for r_idx, rec in enumerate(records, start=2):
            ws.append([rec.get(col) for col, _ in EXPORT_COLS])
            for cell in ws[r_idx]:
                cell.alignment = wrap_align
                cell.border = thin_border

        # Auto-width calculation
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            # Add padding and scaling for proportional fonts, cap at 50 wide
            ws.column_dimensions[col[0].column_letter].width = min((max_len * 1.15) + 4, 50)

        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'form20_export_{ts}.xlsx',
        )

    return jsonify({'error': 'Unsupported format'}), 400


# ── Retro Export (live from AWS RDS: election_result ⋈ election) ─────────────

# Canonical column order — mirrors the historical "SELECT er.*, e.*" shape so the
# exported file stays compatible with the previous RETRO.csv (el_id appears twice).
RETRO_HEADERS = [
    'el_res_id', 'ca_full_name', 'party_abb', 'party_id', 'el_vote_count',
    'el_vote_perc', 'el_rank', 'el_id', 'ac_no', 'state_abb',
    'original_ca_full_name', 'original_party_abb', 'caste', 'caste_category',
    'gender', 'age', 'incumbency', 'el_id', 'el_year', 'el_type',
]

RETRO_SELECT = (
    "SELECT er.el_res_id, er.ca_full_name, er.party_abb, er.party_id, er.el_vote_count, "
    "er.el_vote_perc, er.el_rank, er.el_id, er.ac_no, er.state_abb, "
    "er.original_ca_full_name, er.original_party_abb, er.caste, er.caste_category, "
    "er.gender, er.age, er.incumbency, e.el_id, e.el_year, e.el_type "
    "FROM election_result er JOIN election e ON er.el_id = e.el_id"
)

def fetch_retro_metadata_sync():
    """Build {state: {el_type: {year: distinct_ac_count}}} live from RDS. Full scan — cached."""
    conn = get_rds_db()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT er.state_abb, e.el_type, e.el_year, COUNT(DISTINCT er.ac_no) "
            "FROM election_result er JOIN election e ON er.el_id = e.el_id "
            "GROUP BY er.state_abb, e.el_type, e.el_year"
        )
        meta = {}
        for state, el_type, el_year, cnt in cur.fetchall():
            if state is None or el_type is None or el_year is None:
                continue
            s, t, y = str(state).strip(), str(el_type).strip(), str(el_year).strip()
            meta.setdefault(s, {}).setdefault(t, {})[y] = cnt
        return meta
    except Exception as e:
        print(f"fetch_retro_metadata_sync error: {e}")
        return None
    finally:
        conn.close()


def get_retro_metadata_dict():
    """Return cached retro metadata, falling back to a one-time live build if
    the shared cache is empty (e.g. very first request after a fresh deploy)."""
    cached = cache_get('retro')
    if cached is not None:
        return cached
    meta = fetch_retro_metadata_sync() or {}
    cache_set('retro', meta)
    return meta


def _build_retro():
    """Rebuild the retro metadata cache from RDS. Triggered on retro/dashboard
    reload via trigger_refresh('retro', ...) when the cache is stale — no longer
    polled every 10 minutes."""
    meta = fetch_retro_metadata_sync()
    if meta is not None:
        cache_set('retro', meta)


@app.route('/api/retro/metadata')
def retro_metadata():
    trigger_refresh('retro', _build_retro)   # refresh-on-reload (if stale)
    return jsonify(get_retro_metadata_dict())


@app.route('/api/retro/export')
def export_retro():
    state   = request.args.get('state', '').strip()
    el_type = request.args.get('el_type', '').strip()
    year    = request.args.get('year', '').strip()
    fmt     = request.args.get('format', 'csv').strip().lower()

    if not state or not el_type or not year:
        return jsonify({'error': 'Missing required filters: state, el_type, year'}), 400

    conn = get_rds_db()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cur = conn.cursor()
        query = """
            SELECT er.* 
            FROM election_result er 
            JOIN election e ON er.el_id = e.el_id
            WHERE er.state_abb = %s AND e.el_type = %s AND e.el_year = %s
        """
        cur.execute(query, (state, el_type, year))
        rows = cur.fetchall()
        col_names = [desc[0] for desc in cur.description]
        
        base_name = f"Retro_{state}_{el_type}_{year}"
        
        import io
        if fmt == 'xlsx':
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Retro Data"
            ws.append(col_names)
            for row in rows:
                ws.append(row)
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            
            return send_file(out, download_name=f"{base_name}.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
        else:
            import csv
            from flask import Response
            si = io.StringIO()
            cw = csv.writer(si)
            cw.writerow(col_names)
            cw.writerows(rows)
            
            output = si.getvalue()
            return Response(
                output,
                mimetype="text/csv",
                headers={"Content-Disposition": f"attachment;filename={base_name}.csv"}
            )
            
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/retro/filters')
def get_retro_filters():
    meta = get_retro_metadata_dict()
    state_param   = request.args.get('state', '').strip() or None
    el_type_param = request.args.get('el_type', '').strip() or None

    all_states, all_el_types, all_years = set(), set(), set()
    for s, types in meta.items():
        all_states.add(s)
        for t, years in types.items():
            if state_param is None or s == state_param:
                all_el_types.add(t)
            if (state_param is None or s == state_param) and (el_type_param is None or t == el_type_param):
                all_years.update(years.keys())

    def year_key(y):
        try:
            return int(y)
        except ValueError:
            return 0

    return jsonify({
        'states':   sorted(all_states),
        'el_types': sorted(all_el_types),
        'years':    sorted(all_years, key=year_key),
    }), 200


@app.route('/api/retro/count')
def retro_count():
    state   = request.args.get('state', '').strip()
    el_type = request.args.get('el_type', '').strip()
    year    = request.args.get('year', '').strip()
    if year:
        try:
            yi = int(year)
            if not (1900 <= yi <= 2100):
                raise ValueError
        except ValueError:
            return jsonify({'error': 'Invalid year parameter'}), 400

    meta = get_retro_metadata_dict()
    count = 0
    for s, types in meta.items():
        if state and s != state:
            continue
        for t, years in types.items():
            if el_type and t != el_type:
                continue
            for y, c in years.items():
                if year and y != year:
                    continue
                count += c
    return jsonify({'count': count})


@app.route('/api/reload', methods=['POST'])
def reload_database():
    try:
        from init_db import init_database
        init_database(EXCEL_PATH, DB_PATH)
        return jsonify({'success': True, 'message': 'Database reloaded from Excel'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/api/weekly_momentum')
def weekly_momentum():
    """8-week momentum series + latest-week state table for the Weekly Report dashboard."""
    # ── Weekly snapshots (retro/caste/booth per state per week) ──────────────
    snapshot_path = os.path.join(BASE_DIR, 'weekly_snapshots.json')
    snapshots = {}
    if os.path.exists(snapshot_path):
        try:
            with open(snapshot_path, 'r', encoding='utf-8') as f:
                snapshots = json.load(f)
        except Exception:
            pass

    # ── Form20 weekly election counts (from completion_history) ──────────────
    history = get_completion_history()
    history.pop('_updated', None)

    today_d = datetime.now().date()
    cur_mon  = today_d - timedelta(days=today_d.weekday())

    def _week_mon(date_str):
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
        return (d - timedelta(days=d.weekday())).isoformat()

    f20_by_week = {}
    for key, date_str in history.items():
        try:
            wk = _week_mon(date_str)
            f20_by_week[wk] = f20_by_week.get(wk, 0) + 1
        except Exception:
            pass

    # ── Build week window ─────────────────────────────────────────────────────
    # Default: last 4 weeks (current week + 3 prior)
    # With ?from_week=YYYY-MM-DD&to_week=YYYY-MM-DD → custom range (max 26 weeks)
    from_week_param = request.args.get('from_week', '').strip()
    to_week_param   = request.args.get('to_week',   '').strip()

    if from_week_param:
        try:
            d = datetime.strptime(from_week_param, '%Y-%m-%d').date()
            range_start = d - timedelta(days=d.weekday())   # snap to Monday
        except ValueError:
            range_start = cur_mon - timedelta(days=7 * 3)
    else:
        range_start = cur_mon - timedelta(days=7 * 3)      # default: 4 weeks back

    if to_week_param:
        try:
            d = datetime.strptime(to_week_param, '%Y-%m-%d').date()
            range_end = d - timedelta(days=d.weekday())
        except ValueError:
            range_end = cur_mon
    else:
        range_end = cur_mon

    # Safety cap: never show more than 26 weeks
    MAX_WEEKS = 26
    all_weeks = []
    wk_cursor = range_start
    while wk_cursor <= range_end and len(all_weeks) < MAX_WEEKS:
        all_weeks.append(wk_cursor.isoformat())
        wk_cursor += timedelta(days=7)

    weeks = all_weeks

    def _snap(wk):
        s = snapshots.get(wk, snapshots.get(wk + 'T00:00:00', {}))
        return s if isinstance(s, dict) else {}

    # A snapshot metric may be either a per-state dict ({"TS": 9, ...}) or a
    # pre-summed int total — handle both shapes so the endpoint never errors.
    def _metric_total(val):
        if isinstance(val, dict):
            return sum(v for v in val.values() if isinstance(v, (int, float)))
        if isinstance(val, (int, float)):
            return val
        return 0

    def _metric_dict(val):
        return val if isinstance(val, dict) else {}

    series = []
    for wk in weeks:
        s = _snap(wk)
        series.append({
            'week':        wk,
            'f20':         f20_by_week.get(wk, 0),
            'retro_total': _metric_total(s.get('retro')),
            'caste_total': _metric_total(s.get('caste')),
            'booth_total': _metric_total(s.get('booth')),
        })

    # ── State table: latest available snapshot ────────────────────────────────
    latest_wk = max((k[:10] for k in snapshots), default=None) if snapshots else None
    latest_snap = _snap(latest_wk) if latest_wk else {}
    retro_st = _metric_dict(latest_snap.get('retro'))
    caste_st = _metric_dict(latest_snap.get('caste'))
    booth_st = _metric_dict(latest_snap.get('booth'))

    # Form20 per-state: ACs in Form20 source (from live JSON, approximated by STATE_AC_COUNTS)
    raw_live = _load_json_list(LIVE_EXTRACTED_JSON_PATH)
    f20_by_state = {}
    for item in raw_live:
        st = str(item.get('state', '')).strip()
        et = str(item.get('el_type', '')).strip()
        if st and '-BP' not in et:
            f20_by_state[st] = f20_by_state.get(st, 0) + STATE_AC_COUNTS.get(st, 0)

    # Union of all states with any data
    all_states = sorted(set(list(f20_by_state) + list(retro_st) + list(caste_st) + list(booth_st)))
    state_table = []
    for st in all_states:
        f = f20_by_state.get(st, 0)
        r = retro_st.get(st, 0) if isinstance(retro_st.get(st), (int, float)) else 0
        c = caste_st.get(st, 0) if isinstance(caste_st.get(st), (int, float)) else 0
        b = booth_st.get(st, 0) if isinstance(booth_st.get(st), (int, float)) else 0
        if f + r + c + b == 0:
            continue
        state_table.append({'state': st, 'f20': f, 'retro': r, 'caste': c, 'booth': b})

    # Sort by F20 desc, then retro
    state_table.sort(key=lambda x: (-x['f20'], -x['retro']))

    # ── This-week pushed lists ────────────────────────────────────────────────
    # Target week = ?week=YYYY-MM-DD (any day; snapped to its Monday) or current week.
    req_week = (request.args.get('week') or '').strip()
    if req_week:
        try:
            target_mon = _week_mon(req_week)
        except Exception:
            target_mon = cur_mon.isoformat()
    else:
        target_mon = cur_mon.isoformat()

    # Form 20 (and Retro) — elections completed in the target week.
    # Each completion_history key is "STATE-ELTYPE-YEAR"; AC number ≈ state AC count.
    f20_week = []
    for key, date_str in history.items():
        try:
            if _week_mon(date_str) != target_mon:
                continue
            parts = key.split('-')
            if len(parts) < 3:
                continue
            st = parts[0]
            el_year = parts[-1]
            el_type = '-'.join(parts[1:-1])
            f20_week.append({
                'state':   st,
                'el_type': el_type,
                'el_year': el_year,
                'ac':      STATE_AC_COUNTS.get(st, 0),
            })
        except Exception:
            pass
    # Newest/biggest first for readability
    f20_week.sort(key=lambda x: (-x['ac'], x['state'], x['el_year']))

    # All pushed Form 20 elections (every completed entry, all weeks) for the
    # State-Wise Push Volume table — el_type · el_year · state · AC number.
    f20_all = []
    for key, date_str in history.items():
        parts = key.split('-')
        if len(parts) < 3:
            continue
        st = parts[0]
        el_year = parts[-1]
        el_type = '-'.join(parts[1:-1])
        f20_all.append({
            'state':   st,
            'el_type': el_type,
            'el_year': el_year,
            'ac':      STATE_AC_COUNTS.get(st, 0),
            'date':    date_str,
        })
    # Most recent push first, then larger states
    f20_all.sort(key=lambda x: (x['date'], x['ac']), reverse=True)

    # Booth & Caste — state + AC number for the target week (kept zero for now).
    booth_week = []
    caste_week = []

    return jsonify({
        'series':      series,
        'state_table': state_table,
        'target_week': target_mon,
        'f20_week':    f20_week,
        'f20_all':     f20_all,
        'retro_week':  [],          # zero — Weekly Report tracks pushes, not DB totals
        'booth_week':  booth_week,  # zero — Weekly Report tracks pushes, not DB totals
        'caste_week':  caste_week,  # zero — Weekly Report tracks pushes, not DB totals
    })


@app.route('/api/admin/ac_pct')
def admin_ac_pct():
    """Temporary admin endpoint: AC-wise % per state from form20_summary_view vs ac_election_mapping."""
    data = cache_get('admin_ac_pct')
    if data:
        return jsonify(data)
    return jsonify({'error': 'Cache not ready. Will be populated at midnight.'}), 503


# ── Register Glance Routes ───────────────────────────────────────────────────
# (Moved from glance_routes.py to fix Flask route double-app registry issue)

OTHER_QUERIES = {
    "LGD": """
        SELECT ld.state_abb, 
               0 AS district_count,
               0 AS sub_district_count,
               0 AS village_count,
               COUNT(DISTINCT ld.lgd_code) as lgd_count
        FROM lgd_directory ld GROUP BY ld.state_abb
    """,
    "Ejalshakti": """
        SELECT ep.state_abb, COUNT(ep.lgd_code) AS village_count
        FROM ejalshakti_portal ep
        GROUP BY ep.state_abb
    """,
    "Joshua": """
        SELECT jp.state_abb, COUNT(jp.district_name) AS JOSHUA_district_count
        FROM joshua_population jp GROUP BY jp.state_abb
    """,
    "KYS": """
        SELECT sl.state_abb, COUNT(sl.district_name) AS KYS_distinct_count
        FROM school_locator sl GROUP BY sl.state_abb
    """,
    "Muslim Census": """
        SELECT mc.state_abb, COUNT(mc.district_name) AS MUSLIM_district_count
        FROM muslim_census mc GROUP BY mc.state_abb
    """,
    "SECC": """
        SELECT sa.state_abb, COUNT(sa.lgd_code) AS SECC_lgd_code_count
        FROM secc_abstract sa GROUP BY sa.state_abb
    """
}

def get_other_sources_data(force_refresh=False):
    """Reads other demographic tables from precomputed static JSON.
    These are historical datasets (Census 2011, SECC, etc) that do not change."""
    import json
    import os
    json_path = r'C:\Users\User\.gemini\antigravity\brain\1e9dd4bd-b100-4f6c-bac8-3f5ea7385051\scratch\other_sources.json'
    
    results = {
        "LGD": {}, "Ejalshakti": {}, "Joshua": {},
        "KYS": {}, "Muslim Census": {}, "SECC": {}
    }
    
    if os.path.exists(json_path):
        try:
            with open(json_path, 'r') as f:
                data = json.load(f)
                if data: return data
        except Exception as e:
            pass
            
    return results



@app.route('/state-glance-report')
@login_required
def state_glance_report_page():
    user = session.get('user')
    return render_template('state_glance_report.html', user=user)

@app.route('/api/state_glance/data')
def api_state_glance_data():
    state_abb = request.args.get('state_abb', 'BR').upper().strip()
    
    import os, json, redis
    
    # 1. Try Redis First
    try:
        REDIS_HOST = os.environ.get('REDIS_HOST', 'localhost')
        REDIS_PORT = int(os.environ.get('REDIS_PORT', 6379))
        redis_client = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, db=0, decode_responses=True, socket_connect_timeout=2)
        
        cached_data = redis_client.get(f'state_glance:{state_abb}')
        if cached_data:
            return jsonify({
                "success": True,
                "data": json.loads(cached_data),
                "source": "redis"
            })
    except Exception as e:
        print(f"Redis fetch failed for {state_abb}, falling back to JSON: {e}")
        
    # 2. Fallback to JSON file
    cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'state_glance_cache.json')
    if os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                all_data = json.load(f)
            
            if state_abb in all_data:
                return jsonify({
                    "success": True,
                    "data": all_data[state_abb],
                    "source": "json_fallback"
                })
        except Exception as e:
            print(f"Error reading state_glance_cache.json: {e}")
            
    # 3. Ultimate Fallback if not generated yet or missing
    return jsonify({
        "success": False,
        "data": None,
        "source": "none"
    })

@app.route('/api/country_glance_report/data')
@app.route('/api/country_glance_test/data', endpoint='api_country_glance_test_data')
def api_country_glance_test():
    """Exact data matching the June 15th PDF."""
    force_refresh = request.args.get('refresh', '0') == '1'
    data = cache_get('country_glance_data')
    if force_refresh or not data:
        _build_glance_data(force_refresh=True)
        data = cache_get('country_glance_data')
    if not data:
        cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'glance_cache.json')
        if os.path.exists(cache_file):
            try:
                with open(cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    cache_set('country_glance_data', data)
            except: pass
    if data:
        return jsonify(data)
    else:
        return jsonify({
            "success": True,
            "national_avg": {
                "retro": 0, "form20": 0, "caste": 0, "booth": 0, "muslim": 0,
                "joshua": 0, "kys": 0, "lgd": 0, "ejal": 0, "secc": 0,
                "nrega": 0, "indiastat": 0, "lens": 0
            },
            "matrix": []
        })

@app.route('/api/country_glance/data', endpoint='api_country_glance_data')
def api_country_glance():
    """State-level coverage matrix matching exact PDF business logic."""
    force_refresh = request.args.get('refresh', '0') == '1'
    data = cache_get('country_glance_data')
    if force_refresh or not data:
        _build_glance_data(force_refresh=True)
        data = cache_get('country_glance_data')
    if not data:
        cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'glance_cache.json')
        if os.path.exists(cache_file):
            try:
                with open(cache_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    cache_set('country_glance_data', data)
            except: pass
    if data:
        return jsonify(data)
    else:
        return jsonify({
            "success": True,
            "national_avg": {
                "retro": 0, "form20": 0, "caste": 0, "booth": 0, "muslim": 0,
                "joshua": 0, "kys": 0, "lgd": 0, "ejal": 0, "secc": 0,
                "nrega": 0, "indiastat": 0, "lens": 0
            },
            "matrix": []
        })


@app.route('/api/map/district_data', endpoint='api_map_district_data')
def api_map_district_data():
    """
    Returns district-level coverage data for a given auxiliary metric.
    Query param: metric = ejal | joshua | muslim | secc | kys | lgd
    """
    metric = request.args.get('metric', 'ejal').lower().strip()
    
    cache_key = f'map_district_data_{metric}'
    cached_data = cache_get(cache_key)
    if cached_data:
        return jsonify({
            "success": True,
            "metric":  metric,
            "level":   "district",
            "count":   len(cached_data),
            "data":    cached_data
        })
    else:
        return jsonify({"success": True, "metric": metric, "level": "district", "count": 0, "data": []})

@app.route('/api/map/pc_data', endpoint='api_map_pc_data')
def api_map_pc_data():
    metric = request.args.get('metric', 'retro').lower().strip()
    cache_key = f'map_pc_data_{metric}'
    cached_data = cache_get(cache_key)
    if cached_data is not None:
        return jsonify({
            "success": True, "metric": metric, "level": "pc",
            "count": len(cached_data), "data": cached_data
        })
    return jsonify({"success": True, "metric": metric, "level": "pc", "count": 0, "data": []})

@app.route('/api/map/ac_data', endpoint='api_map_ac_data')
def api_map_ac_data():
    metric = request.args.get('metric', 'retro').lower().strip()
    cache_key = f'map_ac_data_{metric}'
    cached_data = cache_get(cache_key)
    if cached_data is not None:
        return jsonify({
            "success": True, "metric": metric, "level": "ac",
            "count": len(cached_data), "data": cached_data
        })
    return jsonify({"success": True, "metric": metric, "level": "ac", "count": 0, "data": []})


# ── Entry point ──────────────────────────────────────────────────────────────
def _build_glance_data(force_refresh=False):
    import os, json

    # 1️⃣ Check Redis first — it is the primary source of truth
    if not force_refresh:
        cached = cache_get('country_glance_data')
        if cached:
            print("  [cache] Glance data served from Redis.")
            return

    # 2️⃣ Fallback: locally generated computed_glance_data.json
    local_computed = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'computed_glance_data.json')
    if os.path.exists(local_computed):
        try:
            with open(local_computed, 'r') as f:
                out_data = json.load(f)
                cache_set('country_glance_data', out_data)
                print("  [cache] Loaded glance data from computed_glance_data.json -> pushed to Redis.")
                return
        except Exception as e:
            print(f"  [cache] Error reading computed_glance_data.json: {e}")

    # 3️⃣ Fallback: static/data/glance_cache.json
    cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'glance_cache.json')
    if os.path.exists(cache_file):
        try:
            with open(cache_file, 'r') as f:
                out_data = json.load(f)
                cache_set('country_glance_data', out_data)
                print("  [cache] Loaded glance data from glance_cache.json -> pushed to Redis.")
                return
        except Exception as e:
            print(f"  [cache] Error reading glance_cache.json: {e}")

    print("  [warmup] No glance data in Redis or local JSON. Skipping (DB queries disabled).")

def _build_pc_ac_caches(force_refresh=False):
    import os, json

    # 1️⃣ Check Redis first — if all PC/AC metric keys are present, nothing to do
    if not force_refresh:
        redis_keys_present = all(
            cache_get(f'map_pc_data_{m}') is not None or cache_get(f'map_ac_data_{m}') is not None
            for m in ['retro', 'form20']
        )
        if redis_keys_present:
            print("  [cache] PC/AC map caches served from Redis.")
            return

    # 2️⃣ Fallback: load from local JSON and push into Redis
    cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'pc_ac_cache.json')
    if not force_refresh and os.path.exists(cache_file):
        try:
            with open(cache_file, 'r') as f:
                data = json.load(f)
                for k, v in data.items(): cache_set(k, v)
                print("  [warmup] Loaded PC/AC map caches from JSON -> pushed to Redis.")
                return
        except Exception as e:
            print(f"  [cache] Error reading pc_ac_cache.json: {e}")

    print("  [warmup] Building PC/AC map caches...")
    conn = get_rds_db()
    if not conn: return
    try:
        cur = conn.cursor()
        for metric in ['retro', 'form20', 'caste', 'booth']:
            if metric == 'retro':
                retro_meta = cache_get('retro') or {}
                avail_tuples = []
                for st, types in retro_meta.items():
                    for ty, years in (types or {}).items():
                        if '-BP' in str(ty): continue
                        for yr in (years or {}):
                            avail_tuples.append(f"('{str(st).strip()}', '{str(ty).strip()}', {str(yr).strip()})")
                avail_in = ", ".join(avail_tuples) if avail_tuples else "('', '', 0)"
            cov_query = ""
            tot_query = ""
            
            if metric == 'retro':
                retro_meta = cache_get('retro') or {}
                avail_tuples = []
                for st, types in retro_meta.items():
                    for ty, years in (types or {}).items():
                        if '-BP' in str(ty): continue
                        for yr in (years or {}):
                            avail_tuples.append(f"('{str(st).strip()}', '{str(ty).strip()}', {str(yr).strip()})")
                avail_in = ", ".join(avail_tuples) if avail_tuples else "('', '', 0)"
                cov_query = f"SELECT state_abb, ac_no, COUNT(*) as cnt FROM ac_election_mapping WHERE (state_abb, el_type, el_year) IN ({avail_in}) GROUP BY state_abb, ac_no"
                tot_query = "SELECT state_abb, ac_no, COUNT(*) as cnt FROM ac_election_mapping WHERE el_type NOT LIKE '%-BP%' GROUP BY state_abb, ac_no"
            elif metric == 'form20':
                live_meta = cache_get('live') or []
                avail_tuples = []
                for item in live_meta:
                    st, ty, yr = item.get('state'), item.get('el_type'), item.get('el_year')
                    if st and ty and yr and '-BP' not in str(ty):
                        avail_tuples.append(f"('{str(st).strip()}', '{str(ty).strip()}', {str(yr).strip()})")
                avail_in = ", ".join(avail_tuples) if avail_tuples else "('', '', 0)"
                cov_query = f"SELECT state_abb, ac_no, COUNT(*) as cnt FROM ac_election_mapping WHERE (state_abb, el_type, el_year) IN ({avail_in}) GROUP BY state_abb, ac_no"
                tot_query = "SELECT state_abb, ac_no, COUNT(*) as cnt FROM ac_election_mapping WHERE el_type NOT LIKE '%-BP%' GROUP BY state_abb, ac_no"
            elif metric == 'caste':
                cov_query = "SELECT state_abb, ac_no, 1 as cnt FROM caste_details GROUP BY state_abb, ac_no"
                tot_query = "SELECT state_abb, ac_no, 1 as cnt FROM ac_mapping GROUP BY state_abb, ac_no"
            elif metric == 'booth':
                cov_query = "SELECT state_abb, ac_no, 1 as cnt FROM booth_metadata_full_view GROUP BY state_abb, ac_no"
                tot_query = "SELECT state_abb, ac_no, 1 as cnt FROM ac_mapping GROUP BY state_abb, ac_no"
                
            cur.execute(cov_query)
            cov_dict = {f"{r[0]}|{r[1]}": r[2] for r in cur.fetchall()}
            
            cur.execute(tot_query)
            tot_dict = {f"{r[0]}|{r[1]}": r[2] for r in cur.fetchall()}
            
            mapping_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'ac_pc_mapping.json')
            import json
            with open(mapping_file, 'r', encoding='utf-8') as f:
                ac_pc_map = json.load(f)
                
            pc_agg = {}
            ac_data = []
            
            for ac_key, ac_info in ac_pc_map['acs'].items():
                st, ac_no = ac_key.split('|')
                covered = int(cov_dict.get(ac_key, 0))
                total = int(tot_dict.get(ac_key, 1))
                
                ac_data.append({
                    "state_abb": st, "state_name": STATE_NAMES.get(st, st),
                    "ac_no": ac_no, "ac_name": ac_info['ac_name'],
                    "covered": covered, "total": total,
                    "pct": round((covered / total) * 100, 2) if total > 0 else 0
                })
                
                pc_id = ac_info['pc_id']
                if pc_id not in pc_agg:
                    pc_agg[pc_id] = {'covered': 0, 'total': 0}
                pc_agg[pc_id]['covered'] += covered
                pc_agg[pc_id]['total'] += total
                
            pc_data = []
            for pc_id, counts in pc_agg.items():
                if pc_id in ac_pc_map['pcs']:
                    pc_info = ac_pc_map['pcs'][pc_id]
                    covered = counts['covered']
                    total = counts['total']
                    st = pc_info['state_abb']
                    pc_data.append({
                        "state_abb": st, "state_name": STATE_NAMES.get(st, st),
                        "pc_no": pc_info['pc_no'], "pc_name": pc_info['pc_name'],
                        "covered": covered, "total": total,
                        "pct": round((covered / total) * 100, 2) if total > 0 else 0
                    })
                    
            cache_set(f'map_pc_data_{metric}', pc_data)
            cache_set(f'map_ac_data_{metric}', ac_data)

        # BUILD DISTRICT MAP CACHES
        for metric in ['ejal', 'joshua', 'muslim', 'secc', 'kys', 'lgd']:
            data = []
            if metric == 'ejal':
                cur.execute("""
                    WITH ep_agg AS (
                        SELECT DISTINCT lgd_code FROM ejalshakti_portal WHERE lgd_code IS NOT NULL
                    )
                    SELECT
                        ld.state_abb,
                        ld.district_name,
                        COUNT(DISTINCT ep.lgd_code)                    AS covered,
                        COUNT(DISTINCT ld.lgd_code)                    AS total
                    FROM lgd_directory ld
                    LEFT JOIN ep_agg ep ON ep.lgd_code = ld.lgd_code
                    WHERE ld.district_name IS NOT NULL
                    GROUP BY ld.state_abb, ld.district_name
                    ORDER BY ld.state_abb, ld.district_name
                """)
                rows = cur.fetchall()
                for r in rows:
                    state_abb, dist_name, covered, total = r
                    pct = round((covered / total) * 100, 2) if total and total > 0 else 0
                    data.append({
                        "state_abb":     state_abb,
                        "state_name":    STATE_NAMES.get(state_abb, state_abb),
                        "district_name": dist_name,
                        "covered":       covered,
                        "total":         total,
                        "pct":           pct
                    })
            elif metric == 'joshua':
                cur.execute("""
                    SELECT state_abb, district_name, COUNT(DISTINCT id) AS cnt
                    FROM joshua_population
                    WHERE district_name IS NOT NULL
                    GROUP BY state_abb, district_name
                    ORDER BY state_abb, district_name
                """)
                rows = cur.fetchall()
                if not rows:
                    cur.execute("SELECT state_abb, district_name, 0 AS cnt FROM lgd_directory WHERE district_name IS NOT NULL GROUP BY state_abb, district_name")
                    rows = cur.fetchall()
                max_val = max((r[2] for r in rows), default=0)
                for r in rows:
                    state_abb, dist_name, cnt = r
                    data.append({
                        "state_abb":     state_abb,
                        "state_name":    STATE_NAMES.get(state_abb, state_abb),
                        "district_name": dist_name,
                        "covered":       cnt,
                        "total":         max_val,
                        "pct":           round((cnt / max_val) * 100, 2) if max_val > 0 else 0
                    })
            elif metric == 'muslim':
                cur.execute("""
                    SELECT state_abb, district_name, COUNT(DISTINCT id) AS cnt
                    FROM muslim_census
                    WHERE district_name IS NOT NULL
                    GROUP BY state_abb, district_name
                    ORDER BY state_abb, district_name
                """)
                rows = cur.fetchall()
                if not rows:
                    cur.execute("SELECT state_abb, district_name, 0 AS cnt FROM lgd_directory WHERE district_name IS NOT NULL GROUP BY state_abb, district_name")
                    rows = cur.fetchall()
                max_val = max((r[2] for r in rows), default=0)
                for r in rows:
                    state_abb, dist_name, cnt = r
                    data.append({
                        "state_abb":     state_abb,
                        "state_name":    STATE_NAMES.get(state_abb, state_abb),
                        "district_name": dist_name,
                        "covered":       cnt,
                        "total":         max_val,
                        "pct":           round((cnt / max_val) * 100, 2) if max_val > 0 else 0
                    })
            elif metric == 'secc':
                cur.execute("""
                    WITH sa_agg AS (
                        SELECT DISTINCT lgd_code FROM secc_abstract WHERE lgd_code IS NOT NULL
                    )
                    SELECT 
                        ld.state_abb, ld.district_name, 
                        COUNT(DISTINCT sa.lgd_code) AS covered,
                        COUNT(DISTINCT ld.lgd_code) AS total
                    FROM lgd_directory ld
                    LEFT JOIN sa_agg sa ON sa.lgd_code = ld.lgd_code
                    WHERE ld.district_name IS NOT NULL
                    GROUP BY ld.state_abb, ld.district_name
                    ORDER BY ld.state_abb, ld.district_name
                """)
                rows = cur.fetchall()
                for r in rows:
                    state_abb, dist_name, covered, total = r
                    data.append({
                        "state_abb":     state_abb,
                        "state_name":    STATE_NAMES.get(state_abb, state_abb),
                        "district_name": dist_name,
                        "covered":       covered,
                        "total":         total,
                        "pct":           round((covered / total) * 100, 2) if total and total > 0 else 0
                    })
            elif metric == 'kys':
                cur.execute("""
                    SELECT state_abb, district_name, COUNT(DISTINCT id) AS cnt
                    FROM school_locator
                    WHERE district_name IS NOT NULL
                    GROUP BY state_abb, district_name
                    ORDER BY state_abb, district_name
                """)
                rows = cur.fetchall()
                if not rows:
                    cur.execute("SELECT state_abb, district_name, 0 AS cnt FROM lgd_directory WHERE district_name IS NOT NULL GROUP BY state_abb, district_name")
                    rows = cur.fetchall()
                max_val = max((r[2] for r in rows), default=0)
                for r in rows:
                    state_abb, dist_name, cnt = r
                    data.append({
                        "state_abb":     state_abb,
                        "state_name":    STATE_NAMES.get(state_abb, state_abb),
                        "district_name": dist_name,
                        "covered":       cnt,
                        "total":         max_val,
                        "pct":           round((cnt / max_val) * 100, 2) if max_val > 0 else 0
                    })
            elif metric == 'lgd':
                cur.execute("""
                    SELECT state_abb, district_name, COUNT(DISTINCT lgd_code) AS cnt
                    FROM lgd_directory
                    WHERE district_name IS NOT NULL
                    GROUP BY state_abb, district_name
                    ORDER BY state_abb, district_name
                """)
                rows = cur.fetchall()
                max_val = max((r[2] for r in rows), default=0)
                for r in rows:
                    state_abb, dist_name, cnt = r
                    data.append({
                        "state_abb":     state_abb,
                        "state_name":    STATE_NAMES.get(state_abb, state_abb),
                        "district_name": dist_name,
                        "covered":       cnt,
                        "total":         max_val,
                        "pct":           round((cnt / max_val) * 100, 2) if max_val > 0 else 0
                    })
            
            cache_set(f'map_district_data_{metric}', data)
        
        try:
            import os, json
            all_caches = {}
            for metric in ['retro', 'form20', 'caste', 'booth']:
                all_caches[f'map_pc_data_{metric}'] = cache_get(f'map_pc_data_{metric}')
                all_caches[f'map_ac_data_{metric}'] = cache_get(f'map_ac_data_{metric}')
            for metric in ['ejal', 'joshua', 'muslim', 'secc', 'kys', 'lgd']:
                all_caches[f'map_district_data_{metric}'] = cache_get(f'map_district_data_{metric}')
                all_caches[f'map_ac_data_{metric}'] = cache_get(f'map_ac_data_{metric}')
            
            cache_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'pc_ac_cache.json')
            os.makedirs(os.path.dirname(cache_file), exist_ok=True)
            with open(cache_file, 'w') as f: json.dump(all_caches, f)
        except: pass
        
        print("  [warmup] PC/AC map caches ready.")
    except Exception as e:
        print(f"  [warmup] Error building map caches: {e}")
    finally:
        conn.close()

def _refresh_caches_from_db():
    """Force refresh all caches from the database with individual failure handling."""
    print("  [cache] Refreshing caches from DB...")
    
    try:
        analytics = build_analytics_cache(force_refresh=True)
        if analytics:
            cache_set('analytics', analytics)
            print("  [cache] Analytics cache ready.")
        else:
            print("  [cache] Analytics cache build returned empty (no DB connection?).")
    except Exception as e:
        print(f"  [cache] Error building analytics cache: {e}")

    try:
        get_other_sources_data(force_refresh=True)
        print("  [cache] Other sources cache ready.")
    except Exception as e:
        print(f"  [cache] Error building other sources cache: {e}")

    try:
        _build_glance_data(force_refresh=True)
        print("  [cache] Glance data cache ready.")
    except Exception as e:
        print(f"  [cache] Error building glance data cache: {e}")

    try:
        _build_pc_ac_caches(force_refresh=True)
        print("  [cache] PC/AC map caches ready.")
    except Exception as e:
        print(f"  [cache] Error building PC/AC caches: {e}")

    try:
        _build_admin_ac_pct_cache(force_refresh=True)
        print("  [cache] Admin AC Pct cache ready.")
    except Exception as e:
        print(f"  [cache] Error building admin_ac_pct cache: {e}")

    try:
        _build_retro_exports(force_refresh=True)
        print("  [cache] Retro Exports ready.")
    except Exception as e:
        print(f"  [cache] Error building retro exports: {e}")
        
    print("  [cache] Midnight cache refresh process completed.")

def _build_admin_ac_pct_cache(force_refresh=False):
    if not force_refresh: return
    rds = get_rds_db()
    if not rds: return
    try:
        cur = rds.cursor()
        cur.execute("""
            SELECT state_abb, SUM(ac_count) AS total_mapping_acs
            FROM (
                SELECT state_abb, el_type, el_year, COUNT(DISTINCT ac_no) AS ac_count
                FROM ac_election_mapping
                GROUP BY state_abb, el_type, el_year
            ) sub GROUP BY state_abb ORDER BY state_abb
        """)
        mapping = {r[0]: int(r[1]) for r in cur.fetchall()}
        cur.execute("""
            SELECT state_abb, SUM(ac_count) AS total_form20_acs
            FROM (
                SELECT state_abb, el_type, el_year, COUNT(DISTINCT ac_no) AS ac_count
                FROM form20_summary_view
                GROUP BY state_abb, el_type, el_year
            ) sub GROUP BY state_abb ORDER BY state_abb
        """)
        form20 = {r[0]: int(r[1]) for r in cur.fetchall()}
        all_states = sorted(set(list(mapping.keys()) + list(form20.keys())))
        rows = []
        total_f = total_m = 0
        for state in all_states:
            f = form20.get(state, 0); m = mapping.get(state, 0)
            pct = round(f / m * 100, 2) if m > 0 else 0.0
            total_f += f; total_m += m
            rows.append({'state': state, 'form20_acs': f, 'mapping_acs': m, 'pct': pct})
        total_pct = round(total_f / total_m * 100, 2) if total_m > 0 else 0.0
        data = {'rows': rows, 'total_form20': total_f, 'total_mapping': total_m, 'total_pct': total_pct}
        cache_set('admin_ac_pct', data)
    except Exception as e:
        print(f"Error building admin_ac_pct cache: {e}")
    finally:
        rds.close()

def _build_retro_exports(force_refresh=False):
    if not force_refresh: return
    print("  [cache] Building retro export files...")
    meta = fetch_retro_metadata_sync()
    if not meta: return
    conn = get_rds_db()
    if not conn: return
    
    import csv, openpyxl, os
    from openpyxl.styles import Font, PatternFill
    out_dir = os.path.join(os.path.dirname(__file__), 'static', 'data', 'exports')
    os.makedirs(out_dir, exist_ok=True)
    
    try:
        cur = conn.cursor()
        for state, types in meta.items():
            for el_type, years in types.items():
                for year in years.keys():
                    try:
                        year_int = int(year)
                        cur.execute(
                            RETRO_SELECT + " WHERE er.state_abb = %s AND e.el_type = %s AND e.el_year = %s ORDER BY er.ac_no, er.el_rank",
                            (state, el_type, year_int)
                        )
                        rows = cur.fetchall()
                        if not rows: continue
                        
                        base_name = f"Retro_{state}_{el_type}_{year}"
                        csv_path = os.path.join(out_dir, f"{base_name}.csv")
                        xlsx_path = os.path.join(out_dir, f"{base_name}.xlsx")
                        
                        # Build CSV
                        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                            writer = csv.writer(f)
                            writer.writerow(RETRO_HEADERS)
                            writer.writerows(rows)
                            
                        # Build XLSX
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Retro Data"
                        hf = Font(bold=True, color="FFFFFF")
                        fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                        ws.append(RETRO_HEADERS)
                        for cell in ws[1]:
                            cell.font = hf
                            cell.fill = fill
                        for r in rows:
                            ws.append(list(r))
                        ws.freeze_panes = "A2"
                        wb.save(xlsx_path)
                    except Exception as e:
                        print(f"Error building export for {state} {el_type} {year}: {e}")
    finally:
        conn.close()

def _daily_midnight_refresh():
    """Loop forever, waking up exactly at midnight to refresh caches from the DB."""
    import time
    from datetime import datetime, timedelta
    while True:
        now = datetime.now()
        # Calculate time until next midnight
        tomorrow = now + timedelta(days=1)
        midnight = datetime(tomorrow.year, tomorrow.month, tomorrow.day, 0, 0, 0)
        seconds_until_midnight = (midnight - now).total_seconds()
        
        print(f"  [cache-scheduler] Sleeping for {int(seconds_until_midnight)} seconds until midnight...")
        time.sleep(seconds_until_midnight)
        
        print("  [cache-scheduler] Midnight! Triggering DB cache refresh.")
        _refresh_caches_from_db()

def _init_caches_on_startup():
    """On startup: check Redis first. If Redis has all data, serve from it.
    If Redis is cold/empty, fall back to local JSON files and warm Redis.
    NEVER queries the DB during startup — DB is only touched at midnight."""
    import threading

    print("  [startup] Checking Redis for existing cache data...")
    r = get_redis()

    # Check which keys are already live in Redis
    redis_has_analytics       = r and r.exists('cache:analytics:data')
    redis_has_glance          = r and r.exists('cache:country_glance_data:data')
    redis_has_pc_ac           = r and (r.exists('cache:map_pc_data_retro:data') or r.exists('cache:map_ac_data_retro:data'))
    redis_has_live            = r and r.exists('cache:live:data')
    redis_has_retro           = r and r.exists('cache:retro:data')
    redis_has_other_sources   = r and r.exists('cache:other_sources:data')

    all_in_redis = all([
        redis_has_analytics, redis_has_glance, redis_has_pc_ac,
        redis_has_live, redis_has_retro
    ])

    if all_in_redis:
        print("  [startup] ✅ All caches found in Redis — serving directly. No DB, no JSON reads.")
    else:
        print("  [startup] Redis cache is cold or partial. Loading from local JSON files into Redis...")
        # These functions check Redis first, then fall back to JSON, then push to Redis
        build_analytics_cache(force_refresh=False)
        _build_glance_data(force_refresh=False)
        _build_pc_ac_caches(force_refresh=False)
        # Also warm live/retro metadata from JSON if present
        import os, json
        for cache_key, json_path in [
            ('live',  os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'data', 'analytics_cache.json')),
        ]:
            if not (r and r.exists(f'cache:{cache_key}:data')) and os.path.exists(json_path):
                try:
                    with open(json_path) as f:
                        d = json.load(f)
                    if cache_key == 'live' and isinstance(d, dict):
                        live_val = d.get('form20', {}).get('state_progress', [])
                        if live_val:
                            cache_set('live', live_val)
                            print(f"  [startup] Warmed '{cache_key}' cache from JSON.")
                except Exception as e:
                    print(f"  [startup] Could not warm '{cache_key}' from JSON: {e}")
        print("  [startup] JSON -> Redis warmup complete.")

    # Start the midnight scheduler (only refreshes from DB once per night)
    threading.Thread(target=_daily_midnight_refresh, daemon=True).start()
    print("  [startup] Midnight scheduler started.")


if __name__ == '__main__':
    # Local dev: skip Google OAuth so the dashboard works without valid client
    # credentials. Override by exporting DISABLE_AUTH=0 before running.
    os.environ.setdefault('DISABLE_AUTH', '1')

    if not os.path.exists(DB_PATH):
        print("No database found — initialising...")
        try:
            from init_db import init_database
            init_database(EXCEL_PATH, DB_PATH)
        except Exception as e:
            print("Skipping init_db init:", e)

    print("\n  Form 20 Backlog Dashboard")
    print("  Running locally at: http://127.0.0.1:5050")
    if auth_disabled():
        print("  Auth: DISABLED (local dev) — signed in as", DEV_USER['email'])
    print("  Network listening at: http://0.0.0.0:5050\n")

    # Init caches if missing and start the daily midnight refresh schedule
    _init_caches_on_startup()

    app.run(host='0.0.0.0', debug=True, port=5050, use_reloader=False)

# Trigger reload
