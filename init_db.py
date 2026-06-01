"""
init_db.py — Initialize Form 20 Backlog SQLite database from Excel tracker.
Run this directly: python init_db.py
"""

import openpyxl
import sqlite3
import os
from datetime import datetime

STATE_NAMES = {
    'AN': 'Andaman & Nicobar',
    'AP': 'Andhra Pradesh',
    'AR': 'Arunachal Pradesh',
    'AS': 'Assam',
    'BR': 'Bihar',
    'CG': 'Chhattisgarh',
    'CH': 'Chandigarh',
    'DD': 'Daman & Diu',
    'DL': 'Delhi',
    'DN': 'D&NH',
    'GA': 'Goa',
    'GJ': 'Gujarat',
    'HR': 'Haryana',
    'HP': 'Himachal Pradesh',
    'JH': 'Jharkhand',
    'JK': 'Jammu & Kashmir',
    'KA': 'Karnataka',
    'KL': 'Kerala',
    'LA': 'Ladakh',
    'LD': 'Lakshadweep',
    'MH': 'Maharashtra',
    'ML': 'Meghalaya',
    'MN': 'Manipur',
    'MP': 'Madhya Pradesh',
    'MZ': 'Mizoram',
    'NL': 'Nagaland',
    'OR': 'Odisha',
    'PB': 'Punjab',
    'PY': 'Puducherry',
    'RJ': 'Rajasthan',
    'SK': 'Sikkim',
    'TN': 'Tamil Nadu',
    'TR': 'Tripura',
    'TS': 'Telangana',
    'UK': 'Uttarakhand',
    'UP': 'Uttar Pradesh',
    'WB': 'West Bengal',
}


def init_database(excel_path='Form20 Backlog Tracker.xlsx', db_path='data.db'):
    """Read the Excel tracker and build a normalized SQLite database."""

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ── 1. Tracker (Not in DB) ─────────────────────────────────────────────
    ws_tracker = wb['Tracker(Not in DB)']
    tracker_records = []
    tracker_keys = set()
    for row in list(ws_tracker.iter_rows(values_only=True))[1:]:
        state = str(row[0]).strip() if row[0] else None
        el_type = str(row[1]).strip() if row[1] else None
        el_year = int(row[2]) if row[2] else None
        if not (state and el_type and el_year):
            continue
        key = f'{state}-{el_type}-{el_year}'
        sir_flag = (str(row[4]).strip().lower() == 'yes') if row[4] else False
        wip_flag = (str(row[5]).strip().lower() in ('yes', 'y')) if row[5] else False
        remark = str(row[6]).strip() if row[6] else None
        tracker_records.append({
            'state': state,
            'el_type': el_type,
            'el_year': el_year,
            'key': key,
            'sir_state': sir_flag,
            'wip': wip_flag,
            'remark': remark,
        })
        tracker_keys.add(key)

    # ── 2. Download Report CSV — Downloaded records ─────────────────────────────
    import csv
    downloaded_keys = set()
    missing_reasons = {}  # key -> reason string
    csv_path = os.path.join(os.path.dirname(excel_path), 'Form 20 Download Report for Allotted States.csv')
    if os.path.exists(csv_path):
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                state = str(row.get('State Code', '')).strip()
                election = str(row.get('Election', '')).strip()
                year = str(row.get('Year', '')).strip()
                status = str(row.get('Drive Status', '')).strip()
                note = str(row.get('Jira Note', '')).strip()
                if not (state and election and year):
                    continue
                key = f'{state}-{election}-{year}'
                if status == 'Present':
                    downloaded_keys.add(key)
                elif note:
                    missing_reasons[key] = note

    # ── 3. SIR States list ────────────────────────────────────────────────
    ws_sir = wb['SIR States']
    sir_states_set = set()
    for row in list(ws_sir.iter_rows(values_only=True))[1:]:
        if row[0]:
            sir_states_set.add(str(row[0]).strip())

    # ── 4. Form 20 Status — SIR States (extracted/completed) ──────────────
    ws_form20 = wb['Form 20 Status SIR States']
    form20_status = {}
    for row in list(ws_form20.iter_rows(values_only=True))[1:]:
        if row[0] and row[1] and row[2]:
            key = f'{str(row[0]).strip()}-{str(row[1]).strip()}-{int(row[2])}'
            form20_status[key] = {
                'status': row[3],
                'remark': row[4],
            }

    # ── 5. Latest El Year — detailed status per SIR state ─────────────────
    ws_lat = wb['Latest El Year']
    latest_status = {}
    for row in list(ws_lat.iter_rows(values_only=True))[1:]:
        if row[0] and row[1] and row[2]:
            key = f'{str(row[0]).strip()}-{str(row[1]).strip()}-{int(row[2])}'
            latest_status[key] = {
                'avail_status': str(row[3]).strip() if row[3] else None,
                'remark': str(row[4]).strip() if row[4] else None,
                'db': str(row[5]).strip() if row[5] else None,
            }

    # ── Status determination ───────────────────────────────────────────────
    def determine_status(rec):
        key = rec['key']

        # Priority 1: detailed SIR latest-year status
        if key in latest_status:
            lat = latest_status[key]
            db_val = str(lat.get('db') or '').lower()
            remark_val = str(lat.get('remark') or '').lower()
            avail = str(lat.get('avail_status') or '').lower()

            if 'in db' in db_val:
                return 'completed'
            if 'complet' in remark_val:
                return 'extracted'
            if 'not available' in avail:
                return 'missing'
            if key in form20_status:
                return 'extracted'
            return 'downloaded'

        # Priority 2: form20 extraction status for SIR states
        if key in form20_status:
            return 'extracted'

        # Priority 3: WIP
        if rec.get('wip'):
            return 'pending'

        # Priority 4: downloaded (in 25th May list)
        if key in downloaded_keys:
            return 'downloaded'

        return 'missing'

    # ── Build SQLite ───────────────────────────────────────────────────────
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    c.executescript('''
        DROP TABLE IF EXISTS records;
        DROP TABLE IF EXISTS activity_log;

        CREATE TABLE records (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            state               TEXT NOT NULL,
            state_name          TEXT,
            el_type             TEXT NOT NULL,
            el_year             INTEGER NOT NULL,
            key                 TEXT UNIQUE NOT NULL,
            is_sir_state        INTEGER DEFAULT 0,
            download_status     TEXT DEFAULT 'missing',
            extraction_status   TEXT DEFAULT 'pending',
            db_status           TEXT DEFAULT 'not_in_db',
            overall_status      TEXT DEFAULT 'missing',
            wip                 INTEGER DEFAULT 0,
            assigned_to         TEXT,
            remark              TEXT,
            last_updated        TEXT,
            retro_ready         INTEGER DEFAULT 0,
            created_at          TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE activity_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            record_key  TEXT,
            action      TEXT,
            old_value   TEXT,
            new_value   TEXT,
            changed_by  TEXT,
            timestamp   TEXT DEFAULT CURRENT_TIMESTAMP
        );
    ''')

    today = datetime.now().strftime('%Y-%m-%d')
    inserted = 0

    for rec in tracker_records:
        status = determine_status(rec)
        state_name = STATE_NAMES.get(rec['state'], rec['state'])
        download_status = 'downloaded' if rec['key'] in downloaded_keys else 'missing'
        extraction_status = 'extracted' if status in ('extracted', 'completed') else 'pending'
        db_status = 'in_db' if status == 'completed' else 'not_in_db'
        is_sir = 1 if (rec['sir_state'] or rec['state'] in sir_states_set) else 0
        retro_ready = 1 if (is_sir and status in ('extracted', 'completed')) else 0

        # Use CSV missing reason as remark if record has no existing remark
        remark = rec['remark']
        if status == 'missing' and not remark and rec['key'] in missing_reasons:
            remark = missing_reasons[rec['key']]

        c.execute('''
            INSERT OR REPLACE INTO records
                (state, state_name, el_type, el_year, key, is_sir_state,
                 download_status, extraction_status, db_status, overall_status,
                 wip, remark, last_updated, retro_ready, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            rec['state'], state_name, rec['el_type'], rec['el_year'], rec['key'],
            is_sir, download_status, extraction_status, db_status, status,
            1 if rec['wip'] else 0, remark, today, retro_ready,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ))
        inserted += 1

    conn.commit()

    # Summary
    print(f"\n[OK] Inserted {inserted} records into {db_path}")
    print("\nStatus breakdown:")
    rows = c.execute(
        "SELECT overall_status, COUNT(*) FROM records GROUP BY overall_status ORDER BY 2 DESC"
    ).fetchall()
    for status, count in rows:
        print(f"  {status:<12} {count:>4}")

    conn.close()
    print(f"\n[OK] Database ready: {db_path}")


if __name__ == '__main__':
    base = os.path.dirname(os.path.abspath(__file__))
    init_database(
        excel_path=os.path.join(base, 'Form20 Backlog Tracker.xlsx'),
        db_path=os.path.join(base, 'data.db'),
    )
